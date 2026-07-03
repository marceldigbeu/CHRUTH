"""Mission 3 (perf) - ecriture du classeur Prospects_CHAUDS_messages.xlsx
(3 feuilles : Messages, Suivi_Envois, Perf_Messages) avec liste deroulante
sur le statut pour une saisie no-code."""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation

STATUTS = ["A_ENVOYER", "ENVOYE", "REPONDU", "RDV", "REFUS"]


def _ecrire_df(ws, df: pd.DataFrame) -> None:
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(["" if v is None else v for v in r])


def ecrire_classeur(df_messages: pd.DataFrame, df_suivi: pd.DataFrame,
                    df_perf: pd.DataFrame, recommandations: dict,
                    path, seuil: int = 20) -> None:
    wb = Workbook()
    ws_msg = wb.active
    ws_msg.title = "Messages"
    _ecrire_df(ws_msg, df_messages)

    ws_suivi = wb.create_sheet("Suivi_Envois")
    _ecrire_df(ws_suivi, df_suivi)
    # liste deroulante sur la colonne 'statut'
    if "statut" in list(df_suivi.columns):
        col_idx = list(df_suivi.columns).index("statut") + 1
        lettre = get_column_letter(col_idx)
        dv = DataValidation(type="list",
                            formula1='"' + ",".join(STATUTS) + '"', allow_blank=True)
        ws_suivi.add_data_validation(dv)
        dernier = max(ws_suivi.max_row, 2)
        dv.add(f"{lettre}2:{lettre}{dernier}")

    ws_perf = wb.create_sheet("Perf_Messages")
    _ecrire_df(ws_perf, df_perf)
    ws_perf.append([])
    ws_perf.append(["Seuil", seuil, "(resultats min par variante avant bascule)"])
    ws_perf.append(["Variante recommandee par segment"])
    for seg, var in (recommandations or {}).items():
        ws_perf.append([seg, var])

    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
