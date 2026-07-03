from __future__ import annotations

import datetime as _dt

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

NAVY = "1F3864"
WHITE = "FFFFFF"

EURO_FMT = '#,##0" €"'
DATE_FMT = "dd/mm/yyyy"
_NUM_COLS = {"budget_estime_eur"}
_DATE_COLS = {"date_publication", "date_limite"}


def coerce_numeric(value):
    """Renvoie un float si `value` ressemble a un nombre, sinon None."""
    text = str(value or "").strip().replace(" ", "").replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def coerce_date(value):
    """Renvoie une date si `value` est une date ISO/FR, sinon None."""
    text = str(value or "").strip()
    if not text:
        return None
    text = text.split("T")[0].split(" ")[0]
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return _dt.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None

HOT = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
CHECK = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
COLD = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
_HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
_HEADER_FONT = Font(bold=True, color=WHITE)

_INVALID = set(' !@#$%^&*()-+=[]{};:,.<>/?\\|`~"\'')


def _safe_table_name(title: str) -> str:
    name = "tbl_" + "".join("_" if ch in _INVALID else ch for ch in title)
    return name[:255]


def style_data_sheet(ws: Worksheet, freeze_until: str | None = None) -> None:
    if ws.max_row < 1 or ws.max_column < 1:
        return
    headers = {c.value: i for i, c in enumerate(ws[1], start=1)}

    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if ws.max_row >= 2:
        ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        table = Table(displayName=_safe_table_name(ws.title), ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False
        )
        ws.add_table(table)
    else:
        ws.auto_filter.ref = ws.dimensions

    freeze_col = 1
    if freeze_until and freeze_until in headers:
        freeze_col = headers[freeze_until] + 1
    ws.freeze_panes = ws.cell(row=2, column=freeze_col).coordinate

    prio = headers.get("priorite") or headers.get("priorite_excel_formule")
    if prio:
        for row in range(2, ws.max_row + 1):
            val = str(ws.cell(row, prio).value or "").upper()
            fill = HOT if "CHAUD" in val else CHECK if "A_VERIFIER" in val else COLD if "FROID" in val else None
            if fill:
                ws.cell(row, prio).fill = fill

    url = headers.get("url_avis")
    if url:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row, url)
            if cell.value:
                cell.hyperlink = str(cell.value)
                cell.style = "Hyperlink"

    # Formats nombres / dates (les valeurs SQLite arrivent en texte : on coerce)
    for name, col in headers.items():
        if name in _NUM_COLS:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row, col)
                num = coerce_numeric(cell.value)
                if num is not None:
                    cell.value = num
                    cell.number_format = EURO_FMT
        elif name in _DATE_COLS:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row, col)
                parsed = coerce_date(cell.value)
                if parsed is not None:
                    cell.value = parsed
                    cell.number_format = DATE_FMT

    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row in range(1, min(ws.max_row, 60) + 1):
            value = ws.cell(row, col_idx).value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 55)
