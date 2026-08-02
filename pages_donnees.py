"""Base de donnees et fichiers metier dans la plateforme CHRUTH."""
from __future__ import annotations

from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

import liens_source
import livrables_chruth as livrables
import provenance_ao

try:
    st.set_page_config(page_title="Base de donnees CHRUTH", layout="wide")
except st.errors.StreamlitAPIException:
    pass


@st.cache_data(show_spinner=False)
def onglets_excel(chemin: str, _modifie: int) -> list[str]:
    wb = load_workbook(chemin, read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


@st.cache_data(show_spinner=False)
def lire_onglet(chemin: str, _modifie: int, onglet: str) -> pd.DataFrame:
    return pd.read_excel(chemin, sheet_name=onglet, engine="openpyxl")


@st.cache_data(show_spinner=False)
def filtres_prospects(chemin: str, _modifie: int) -> tuple[list[str], list[str], list[str], int]:
    domaines: set[str] = set()
    departements: set[str] = set()
    categories: set[str] = set()
    total = 0
    colonnes = ["domaine_chruth", "code_departement", "categorie_chruth"]
    for bloc in pd.read_csv(chemin, usecols=colonnes, dtype=str, keep_default_na=False,
                            chunksize=20_000, low_memory=False):
        total += len(bloc)
        domaines.update(v for v in bloc["domaine_chruth"].unique() if v)
        departements.update(v for v in bloc["code_departement"].unique() if v)
        categories.update(v for v in bloc["categorie_chruth"].unique() if v)
    return sorted(domaines), sorted(departements), sorted(categories), total


@st.cache_data(show_spinner=False)
def chercher_prospects(chemin: str, _modifie: int, recherche: str,
                       domaines: tuple[str, ...], departements: tuple[str, ...],
                       categories: tuple[str, ...], limite: int) -> tuple[pd.DataFrame, int]:
    colonnes = [
        "siret", "denomination", "enseigne", "categorie_chruth", "domaine_chruth",
        "adresse_complete", "libelle_commune", "code_departement", "effectif_label",
        "latitude", "longitude",
    ]
    trouves: list[pd.DataFrame] = []
    nb_trouves = 0
    terme = recherche.strip().casefold()
    for bloc in pd.read_csv(chemin, usecols=colonnes, dtype=str, keep_default_na=False,
                            chunksize=20_000, low_memory=False):
        masque = pd.Series(True, index=bloc.index)
        if domaines:
            masque &= bloc["domaine_chruth"].isin(domaines)
        if departements:
            masque &= bloc["code_departement"].isin(departements)
        if categories:
            masque &= bloc["categorie_chruth"].isin(categories)
        if terme:
            texte = (bloc["denomination"] + " " + bloc["enseigne"] + " "
                     + bloc["adresse_complete"] + " " + bloc["siret"]).str.casefold()
            masque &= texte.str.contains(terme, regex=False)
        selection = bloc.loc[masque]
        nb_trouves += len(selection)
        deja = sum(len(df) for df in trouves)
        if deja < limite and not selection.empty:
            trouves.append(selection.head(limite - deja))
    resultat = pd.concat(trouves, ignore_index=True) if trouves else pd.DataFrame(columns=colonnes)
    return resultat, nb_trouves


def filtrer_table(df: pd.DataFrame, recherche: str) -> pd.DataFrame:
    terme = recherche.strip().casefold()
    if not terme or df.empty:
        return df
    texte = df.fillna("").astype(str).agg(" ".join, axis=1).str.casefold()
    return df.loc[texte.str.contains(terme, regex=False)]


st.title("Base de donnees et fichiers")
source = livrables.dossier_output()
st.caption(f"Source active : `{source}`")

tab_ao, tab_prospects, tab_fichiers = st.tabs(
    ["Appels d'offres", "Prospects", "Fichiers importants"])

with tab_ao:
    classeur = livrables.fichier("AO_CHRUTH.xlsm")
    if not classeur.exists():
        st.error(f"Classeur introuvable : {classeur}")
    else:
        modifie = classeur.stat().st_mtime_ns
        feuilles = onglets_excel(str(classeur), modifie)
        suggerees = [n for n in ("AO_Tous", "AO_CHAUDS", "AO_A_VERIFIER", "CRM_Suivi")
                     if n in feuilles]
        disponibles = suggerees + [n for n in feuilles if n not in suggerees]
        c1, c2 = st.columns([1, 2])
        onglet = c1.selectbox("Onglet Excel", disponibles)
        recherche_ao = c2.text_input("Rechercher dans les appels d'offres",
                                     placeholder="Objet, acheteur, ville, identifiant...")
        with st.spinner("Lecture du classeur Excel..."):
            table = lire_onglet(str(classeur), modifie, onglet)
        table = filtrer_table(table, recherche_ao)
        if not table.empty and ("source" in table.columns or "url_avis" in table.columns):
            provenances = table.apply(
                lambda ligne: provenance_ao.detecter(ligne.to_dict()), axis=1)
            table = table.copy()
            table["Découvert via"] = [p[0] for p in provenances]
            table["Plateforme du dossier"] = [p[1] for p in provenances]
            ordre = ["Découvert via", "Plateforme du dossier"]
            table = table[ordre + [c for c in table.columns if c not in ordre]]
        m1, m2, m3 = st.columns(3)
        m1.metric("Lignes affichees", len(table))
        m2.metric("Colonnes", len(table.columns))
        m3.metric("Onglet", onglet)
        # Les colonnes d'adresse deviennent cliquables : sans cela, remonter a
        # l'avis d'origine imposait de copier l'URL a la main hors de l'app.
        config = {c: st.column_config.LinkColumn(liens_source.libelle(c),
                                                 display_text="Ouvrir")
                  for c in liens_source.colonnes_de_lien(table.columns)}
        st.dataframe(table, width="stretch", height=520, hide_index=True,
                     column_config=config)
        st.download_button("Exporter la vue en CSV", table.to_csv(index=False).encode("utf-8-sig"),
                           file_name=f"{onglet}.csv", mime="text/csv")

with tab_prospects:
    base = livrables.fichier("prospects_enrichis.csv")
    if not base.exists():
        st.error(f"Base prospects introuvable : {base}")
    else:
        modifie = base.stat().st_mtime_ns
        with st.spinner("Indexation des filtres de la base prospects..."):
            domaines, deps, categories, total = filtres_prospects(str(base), modifie)
        c1, c2, c3 = st.columns(3)
        choix_domaines = c1.multiselect("Domaine", domaines)
        choix_deps = c2.multiselect("Departement", deps)
        choix_categories = c3.multiselect("Categorie", categories)
        recherche = st.text_input("Rechercher un prospect",
                                   placeholder="Societe, enseigne, adresse ou SIRET")
        limite = st.select_slider("Nombre maximal de lignes affichees",
                                  options=[100, 250, 500, 1000, 2000], value=500)
        with st.spinner("Filtrage de la base de 132 000 prospects..."):
            resultat, nb_trouves = chercher_prospects(
                str(base), modifie, recherche, tuple(choix_domaines), tuple(choix_deps),
                tuple(choix_categories), limite)
        m1, m2, m3 = st.columns(3)
        m1.metric("Prospects dans la base", f"{total:,}".replace(",", " "))
        m2.metric("Resultats", f"{nb_trouves:,}".replace(",", " "))
        m3.metric("Affiches", len(resultat))
        st.dataframe(resultat, width="stretch", height=520, hide_index=True)
        st.download_button("Exporter les resultats affiches", 
                           resultat.to_csv(index=False).encode("utf-8-sig"),
                           file_name="prospects_filtres.csv", mime="text/csv")

with tab_fichiers:
    lignes = []
    for libelle, nom in livrables.FICHIERS_IMPORTANTS.items():
        path = livrables.fichier(nom)
        info = livrables.informations(path)
        lignes.append({
            "Fichier": libelle,
            "Nom": nom,
            "Disponible": "Oui" if info["existe"] else "Non",
            "Taille": livrables.taille_humaine(int(info["taille"])),
            "Modifie": info["modifie"],
        })
    st.dataframe(pd.DataFrame(lignes), width="stretch", hide_index=True)

    noms_disponibles = {libelle: nom for libelle, nom in livrables.FICHIERS_IMPORTANTS.items()
                        if livrables.fichier(nom).is_file()}
    if noms_disponibles:
        choix = st.selectbox("Fichier a telecharger", list(noms_disponibles))
        path = livrables.fichier(noms_disponibles[choix])
        taille = path.stat().st_size
        if taille > 25 * 1024 * 1024:
            st.warning("Fichier volumineux : sa preparation peut prendre quelques secondes.")
        cle = "fichier_prepare_" + path.name
        if st.button("Preparer le telechargement", key="preparer_fichier"):
            st.session_state[cle] = path.read_bytes()
        if cle in st.session_state:
            st.download_button("Telecharger", st.session_state[cle], file_name=path.name,
                               mime="application/octet-stream", key="telecharger_fichier")
