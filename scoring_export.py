import os
import pandas as pd
from pathlib import Path
from datetime import datetime

from config import (
    poids_effectif, POIDS_CATEGORIE, SEUILS_SCORE,
    distance_au_siege, poids_proximite,
)

OUTPUT_DIR = Path(__file__).parent / "output"
OUTPUT_DIR.mkdir(exist_ok=True)

# Au-dela de ce nombre de lignes, on n'applique pas la coloration ligne a ligne
# (en-tete + filtre suffisent) : eviter de parcourir des feuilles de 100k+ lignes
# qui faisaient exploser la memoire / le temps a l'export.
COLOR_ROW_CAP = 5000


def load_enriched() -> pd.DataFrame:
    path = OUTPUT_DIR / "prospects_enrichis.csv"
    if not path.exists():
        path = OUTPUT_DIR / "prospects_nettoyes.csv"
    if not path.exists():
        raise FileNotFoundError(
            f"Aucun fichier prospects trouvé dans {OUTPUT_DIR}. "
            "Exécute d'abord clean_classify.py"
        )
    df = pd.read_csv(path, dtype=str)
    print(f"Chargé : {len(df)} lignes depuis {path.name}")
    return df


def to_float(val) -> float:
    try:
        return float(val) if pd.notna(val) else 0.0
    except (ValueError, TypeError):
        return 0.0


def calculer_signal_besoin(row: dict) -> int:
    score = 0

    # Taille de la structure : plus il y a de salaries, plus le besoin de
    # nettoyage/entretien est important.
    score += poids_effectif(int(to_float(row.get("effectif_nombre", 0))))

    # Type de structure (hopital, ecole, mairie... = besoin recurrent fort).
    categorie = str(row.get("categorie_chruth", ""))
    score += POIDS_CATEGORIE.get(categorie, 5)

    # Proximite au siege CHRUTH (Paris 8e) : capacite a servir le client.
    # Remplace l'ancien bonus region flou (+5) par des bandes de distance.
    km = distance_au_siege(row.get("latitude"), row.get("longitude"))
    score += poids_proximite(km)

    return min(score, 100)


def priorite(score: float) -> str:
    if score >= SEUILS_SCORE["CHAUDE"]:
        return "CHAUDE"
    elif score >= SEUILS_SCORE["TIEDE"]:
        return "TIEDE"
    return "FROIDE"


def domaine_from_categorie(categorie: object) -> str:
    value = str(categorie or "").strip().upper()
    if value.startswith("PUB_"):
        return "PUBLIC"
    if value in {"", "AUTRE", "NAN", "NONE"}:
        return "A_CLASSER"
    return "PRIVE"


def ensure_business_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    if "domaine_chruth" not in df.columns:
        df["domaine_chruth"] = df["categorie_chruth"].map(domaine_from_categorie)
    else:
        domaine = df["domaine_chruth"].fillna("").astype(str).str.strip().str.upper()
        inferred = df["categorie_chruth"].map(domaine_from_categorie)
        df["domaine_chruth"] = domaine.where(domaine != "", inferred)

    if "adresse_complete" not in df.columns:
        parts = []
        for column in ["numero_voie", "type_voie", "libelle_voie"]:
            if column in df.columns:
                parts.append(df[column].fillna("").astype(str))
        df["adresse_complete"] = parts[0] if parts else ""
        for part in parts[1:]:
            df["adresse_complete"] = (df["adresse_complete"] + " " + part).str.strip()

    for column in ["libelle_commune", "code_departement", "denomination", "siren", "siret"]:
        if column not in df.columns:
            df[column] = ""

    return df


def first_non_empty(values: pd.Series) -> str:
    for value in values:
        text = str(value or "").strip()
        if text and text.lower() not in {"nan", "none"}:
            return text
    return ""


def unique_join(values: pd.Series, limit: int = 8) -> str:
    seen = []
    for value in values:
        text = str(value or "").strip()
        if not text or text.lower() in {"nan", "none"}:
            continue
        if text not in seen:
            seen.append(text)
        if len(seen) >= limit:
            break
    return ", ".join(seen)


def best_priority(values: pd.Series) -> str:
    rank = {"CHAUDE": 3, "TIEDE": 2, "FROIDE": 1}
    best = ""
    best_rank = -1
    for value in values:
        text = str(value or "").strip().upper()
        if rank.get(text, 0) > best_rank:
            best = text
            best_rank = rank.get(text, 0)
    return best


def build_villes(df: pd.DataFrame) -> pd.DataFrame:
    group_cols = ["domaine_chruth", "code_departement", "libelle_commune"]
    villes = (
        df.groupby(group_cols, dropna=False)
        .agg(
            nb_prospects=("siret", "count"),
            nb_societes=("siren", "nunique"),
            nb_chaude=("priorite", lambda values: int((values == "CHAUDE").sum())),
            score_moyen=("signal_besoin", "mean"),
            categories_principales=("categorie_chruth", unique_join),
        )
        .reset_index()
        .sort_values(["nb_chaude", "score_moyen", "nb_prospects"], ascending=False)
    )
    villes.insert(0, "rang_ville", range(1, len(villes) + 1))
    villes["score_moyen"] = villes["score_moyen"].round(1)
    return villes


def build_adresses(df: pd.DataFrame) -> pd.DataFrame:
    columns = [
        "domaine_chruth", "code_departement", "libelle_commune", "adresse_complete",
        "denomination", "siret", "siren", "categorie_chruth", "sous_categorie",
        "signal_besoin", "priorite", "telephone_finess", "site_web", "latitude", "longitude",
    ]
    existing = [column for column in columns if column in df.columns]
    adresses = df[existing].copy()
    adresses = adresses.sort_values(
        ["domaine_chruth", "code_departement", "libelle_commune", "signal_besoin"],
        ascending=[True, True, True, False],
    )
    if "siret" in adresses.columns:
        adresses = adresses.drop_duplicates(subset=["siret"], keep="first")
    adresses.insert(0, "rang_adresse", range(1, len(adresses) + 1))
    return adresses


def build_societes(df: pd.DataFrame) -> pd.DataFrame:
    work = df.copy()
    work["_identifiant_societe"] = work["siren"].fillna("").astype(str).str.strip()
    missing = work["_identifiant_societe"].isin(["", "nan", "None"])
    work.loc[missing, "_identifiant_societe"] = work.loc[missing, "siret"].fillna("").astype(str).str.strip()

    societes = (
        work.groupby("_identifiant_societe", dropna=False)
        .agg(
            denomination=("denomination", first_non_empty),
            domaines=("domaine_chruth", unique_join),
            categories=("categorie_chruth", unique_join),
            nb_etablissements=("siret", "count"),
            nb_villes=("libelle_commune", "nunique"),
            villes=("libelle_commune", unique_join),
            departements=("code_departement", unique_join),
            adresses=("adresse_complete", unique_join),
            meilleure_priorite=("priorite", best_priority),
            score_max=("signal_besoin", "max"),
            score_moyen=("signal_besoin", "mean"),
        )
        .reset_index()
        .rename(columns={"_identifiant_societe": "identifiant_societe"})
        .sort_values(["score_max", "nb_etablissements"], ascending=False)
    )
    societes.insert(0, "rang_societe", range(1, len(societes) + 1))
    societes["score_moyen"] = societes["score_moyen"].round(1)
    return societes


def build_veille() -> pd.DataFrame:
    rows = [
        {
            "bloc_veille": "Mise a jour hebdomadaire",
            "objectif": "Garder la base proche de la realite terrain.",
            "source": "API Entreprises + pipeline CHRUTH",
            "frequence": "Chaque semaine",
            "action_precise": "Lancer la mise a jour, comparer les nouveaux prospects et verifier les sorties du Top_Cibles.",
            "sortie_attendue": "Nouveau fichier Excel, KPI, controle qualite, exports Power BI et Notion.",
        },
        {
            "bloc_veille": "Domaine public",
            "objectif": "Identifier les structures publiques et les opportunites MAPA.",
            "source": "Domaine_PUBLIC, MAPA_Prioritaires, BOAMP, marches publics locaux",
            "frequence": "Chaque semaine",
            "action_precise": "Filtrer PUBLIC, priorite CHAUDE/TIEDE, puis rechercher les signaux d'appels d'offres ou de renouvellement.",
            "sortie_attendue": "Liste courte de comptes publics a suivre.",
        },
        {
            "bloc_veille": "Domaine prive",
            "objectif": "Prioriser les entreprises privees les plus activables commercialement.",
            "source": "Domaine_PRIVE, CHAUDE, Societes",
            "frequence": "Chaque semaine",
            "action_precise": "Isoler les societes privees CHAUDE, regrouper par ville et preparer les messages de prospection.",
            "sortie_attendue": "CRM_CHRUTH_CHAUDE.xlsx et messages de prospection.",
        },
        {
            "bloc_veille": "Villes et adresses",
            "objectif": "Savoir ou concentrer les actions terrain et verifier la qualite des coordonnees.",
            "source": "Villes et Adresses",
            "frequence": "Chaque semaine",
            "action_precise": "Controler les villes avec beaucoup de prospects CHAUDE, puis verifier les adresses vides ou douteuses.",
            "sortie_attendue": "Priorites geographiques et corrections de donnees.",
        },
        {
            "bloc_veille": "Enumeration des societes",
            "objectif": "Eviter de contacter plusieurs fois la meme societe et reperer les groupes multi-etablissements.",
            "source": "Societes",
            "frequence": "Chaque semaine",
            "action_precise": "Traiter les societes par rang, score maximum et nombre d'etablissements.",
            "sortie_attendue": "Plan de contact par societe plutot que par ligne SIRET uniquement.",
        },
        {
            "bloc_veille": "Sante et medico-social",
            "objectif": "Ameliorer les informations FINESS et les segments sante.",
            "source": "FINESS / data.gouv.fr + enrich_finess.py",
            "frequence": "Chaque mois",
            "action_precise": "Verifier les structures SANTE et les correspondances FINESS lorsque les donnees sont incompletes.",
            "sortie_attendue": "Contacts sante plus fiables.",
        },
        {
            "bloc_veille": "Pilotage Power BI",
            "objectif": "Suivre l'evolution des KPI et des performances de prospection.",
            "source": "output/powerbi_sources",
            "frequence": "Apres chaque mise a jour",
            "action_precise": "Rafraichir Power BI et suivre les volumes par domaine, ville, departement, categorie et priorite.",
            "sortie_attendue": "Tableau de bord a jour.",
        },
    ]
    return pd.DataFrame(rows)


def nom_fichier_sortie():
    return OUTPUT_DIR / "Base_Prospects_CHRUTH.xlsm"


def _messages_actives() -> bool:
    return os.environ.get("CHRUTH_GENERER_MESSAGES", "").strip() in ("1", "true", "True")


def sync_collecte_sheet(wb) -> None:
    from openpyxl.styles import PatternFill

    from ao_config import collecte_active

    ws = wb["MAJ"] if "MAJ" in wb.sheetnames else wb.create_sheet("MAJ", 0)
    actif = collecte_active()
    ws["A1"] = "CHRUTH - Prospects : cliquer le bouton pour mettre a jour"
    ws["A3"] = "Collecte donnees"
    ws["B3"] = "ON" if actif else "OFF"
    ws["B3"].fill = PatternFill("solid", fgColor="C6EFCE" if actif else "FFC7CE")
    ws.column_dimensions["A"].width = 58
    ws.column_dimensions["B"].width = 16


def exporter_excel(df: pd.DataFrame, filepath: Path):
    # Mission 4 : rentabilite ESTIMEE (modele a priori) -> colonnes ca_estime_eur /
    # marge_estimee_eur sur chaque prospect + onglet agrege par segment.
    import rentabilite
    df = rentabilite.enrichir(df)
    table_rentabilite = rentabilite.table_segments(df)


    hot = df[df["priorite"] == "CHAUDE"].sort_values("signal_besoin", ascending=False)
    warm = df[df["priorite"] == "TIEDE"].sort_values("signal_besoin", ascending=False)
    cold = df[df["priorite"] == "FROIDE"].sort_values("signal_besoin", ascending=False)
    public = df[df["domaine_chruth"] == "PUBLIC"].sort_values("signal_besoin", ascending=False)
    prive = df[df["domaine_chruth"] == "PRIVE"].sort_values("signal_besoin", ascending=False)
    a_classer = df[df["domaine_chruth"] == "A_CLASSER"].sort_values("signal_besoin", ascending=False)
    top100 = df.sort_values("signal_besoin", ascending=False).head(100)

    # Mission 3 : brouillons de prospection par segment, uniquement sur les
    # sous-ensembles actionnables (CHAUDE, Top_Cibles) -- jamais sur df (132k).
    table_templates = None
    if _messages_actives():
        try:
            import prospect_messages as pmod
            _tpl = pmod.generer_templates(pmod.segments_activables(df))
            hot = pmod.appliquer_sur(hot, _tpl)
            top100 = pmod.appliquer_sur(top100, _tpl)
            table_templates = pmod.table_templates(_tpl)
        except Exception as exc:  # noqa: BLE001
            print(f"[MESSAGES] generation ignoree : {exc}")

    villes = build_villes(df)
    adresses = build_adresses(df)
    societes = build_societes(df)
    veille = build_veille()

    stats = (
        df.groupby(["domaine_chruth", "categorie_chruth"], dropna=False)
        .agg(
            nb_prospects=("siret", "count"),
            nb_societes=("siren", "nunique"),
            nb_chaude=("priorite", lambda values: int((values == "CHAUDE").sum())),
            score_moyen=("signal_besoin", "mean"),
        )
        .reset_index()
        .sort_values(["domaine_chruth", "nb_prospects"], ascending=[True, False])
    )
    stats["score_moyen"] = stats["score_moyen"].round(1)

    synopsis = pd.DataFrame({
        "Indicateur": [
            "Date d'export",
            "Total prospects",
            "Dont CHAUDE",
            "Dont TIEDE",
            "Dont FROIDE",
            "Dont domaine PUBLIC",
            "Dont domaine PRIVE",
            "Dont a classer",
            "Villes distinctes",
            "Societes distinctes",
            "Top 1 departement",
            "Top 1 ville",
            "Top 1 categorie",
        ],
        "Valeur": [
            datetime.now().strftime("%d/%m/%Y %H:%M"),
            len(df),
            len(hot),
            len(warm),
            len(cold),
            len(public),
            len(prive),
            len(a_classer),
            df["libelle_commune"].nunique(),
            df["siren"].nunique(),
            df["code_departement"].mode().iloc[0] if not df.empty else "",
            df["libelle_commune"].mode().iloc[0] if not df.empty else "",
            df["categorie_chruth"].mode().iloc[0] if not df.empty else "",
        ],
    })

    from openpyxl import load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    feuilles = {
        "Synthese": synopsis, "Prospects": df, "Top_Cibles": top100, "CHAUDE": hot,
        "Domaine_PUBLIC": public, "Domaine_PRIVE": prive, "A_CLASSER": a_classer,
        "MAPA_Prioritaires": public, "Villes": villes, "Adresses": adresses,
        "Societes": societes, "Veille": veille, "Stats": stats,
        "Rentabilite_Estimee": table_rentabilite,
        **({"Templates_Segments": table_templates} if table_templates is not None else {}),
    }
    template = Path(__file__).parent / "assets" / "PROSPECTS_TEMPLATE.xlsm"
    wb = load_workbook(template, keep_vba=True)
    sync_collecte_sheet(wb)
    for nom, frame in feuilles.items():
        ws = wb[nom] if nom in wb.sheetnames else wb.create_sheet(nom)
        if ws.max_row:
            ws.delete_rows(1, ws.max_row)
        for r in dataframe_to_rows(frame, index=False, header=True):
            ws.append(["" if v is None else v for v in r])
    # Mise en forme EN MEMOIRE (une seule passe, pas de rechargement du fichier).
    styler_workbook(wb)
    # Modele financier INTERACTIF (formules) : feuille dediee ecrite apres le style.
    import previsions
    ws_fin = (wb["Modele_Financier"] if "Modele_Financier" in wb.sheetnames
              else wb.create_sheet("Modele_Financier"))
    previsions.ecrire_feuille(ws_fin)
    # Simulateur de rentabilite par marche (Sonia 2026-06-23) : 4 feuilles dediees.
    import rentabilite_marche
    rentabilite_marche.ecrire_feuilles(wb)
    wb.save(filepath)

    print(f"\nExporte : {filepath}")
    print(
        "  Onglets : Synthese, Prospects, Top_Cibles, CHAUDE, Domaine_PUBLIC, "
        "Domaine_PRIVE, A_CLASSER, MAPA_Prioritaires, Villes, Adresses, Societes, Veille, Stats"
    )


def styler_workbook(wb):
    """Mise en forme appliquee EN MEMOIRE (pas de rechargement du fichier).

    En-tete + filtre + largeurs sur toutes les feuilles (peu couteux). La
    coloration ligne a ligne (priorite/score) n'est appliquee que sur les
    feuilles <= COLOR_ROW_CAP lignes : sur les feuilles geantes (Prospects,
    Adresses, Domaine_*...) elle ferait exploser memoire/temps a l'export.
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    fill_chaude = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_tiede = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    fill_froide = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for ws in wb.worksheets:
        if ws.title in ("Synthese", "Synthèse"):
            continue

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        ws.auto_filter.ref = ws.dimensions

        # Coloration ligne a ligne seulement sur les feuilles raisonnables.
        if ws.max_row <= COLOR_ROW_CAP:
            col_priorite = None
            col_score = None
            for idx, cell in enumerate(ws[1], start=1):
                if cell.value == "priorite":
                    col_priorite = idx
                if cell.value == "signal_besoin":
                    col_score = idx

            if col_priorite:
                fills = {"CHAUDE": fill_chaude, "TIEDE": fill_tiede, "FROIDE": fill_froide}
                center = Alignment(horizontal="center")
                for row in ws.iter_rows(min_row=2, max_col=col_priorite, max_row=ws.max_row):
                    cell = row[col_priorite - 1]
                    val = str(cell.value).strip().upper() if cell.value else ""
                    if val in fills:
                        cell.fill = fills[val]
                    cell.alignment = center

            if col_score:
                center = Alignment(horizontal="center")
                for row in ws.iter_rows(min_row=2, max_col=col_score, max_row=ws.max_row):
                    row[col_score - 1].alignment = center

        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row in ws.iter_rows(min_row=1, max_col=col_idx, max_row=min(ws.max_row, 50)):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 3, 50)


def main():
    print("=== SCORING & EXPORT EXCEL ===")

    df = load_enriched()
    df = ensure_business_columns(df)
    df["signal_besoin"] = df.apply(calculer_signal_besoin, axis=1)
    df["priorite"] = df["signal_besoin"].apply(priorite)

    print("\nRépartition des priorités :")
    print(df["priorite"].value_counts().to_string())
    print(f"\nScore moyen : {df['signal_besoin'].mean():.1f}")
    print(f"Score max : {df['signal_besoin'].max()}")

    filepath = nom_fichier_sortie()
    exporter_excel(df, filepath)

    print("\n=== PIPELINE TERMINÉ AVEC SUCCÈS ===")
    print(f"Fichier final : {filepath}")
    print("\nPour l'équipe commerciale : ouvrir l'onglet 'CHAUDE' ou 'Top_Cibles'")


if __name__ == "__main__":
    main()
