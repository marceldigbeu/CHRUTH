"""Modele financier previsionnel CHRUTH (Mission - hypotheses, SANS donnees reelles).

Transforme l'effort de prospection en projection : entonnoir
(contacts -> conversion -> clients -> CA) puis marge, couts fixes, point mort.
Toutes les entrees sont des HYPOTHESES ajustables (config.PREV_*), a valider avec
l'entreprise. Le CRM de capture fournira progressivement les valeurs reelles
(taux de conversion, panier) pour calibrer ce modele.
"""
from __future__ import annotations

import pandas as pd

import config


def parametres() -> dict:
    return {
        "horizon_mois": config.PREV_HORIZON_MOIS,
        "contacts_par_mois": config.PREV_CONTACTS_PAR_MOIS,
        "taux_conversion": config.PREV_TAUX_CONVERSION,
        "panier_annuel": config.PREV_PANIER_MOYEN_ANNUEL_EUR,
        "cout_variable_pct": config.PREV_COUT_VARIABLE_PCT,
        "couts_fixes_mensuels": config.PREV_COUTS_FIXES_MENSUELS_EUR,
        "churn_mensuel": config.PREV_CHURN_MENSUEL,
    }


def projection(params: dict | None = None) -> pd.DataFrame:
    p = params or parametres()
    panier_mensuel = p["panier_annuel"] / 12.0
    clients, cumul, rows = 0.0, 0.0, []
    for m in range(1, int(p["horizon_mois"]) + 1):
        nouveaux = p["contacts_par_mois"] * p["taux_conversion"]
        clients = clients * (1 - p["churn_mensuel"]) + nouveaux
        ca = clients * panier_mensuel
        cout_var = ca * p["cout_variable_pct"]
        marge = ca - cout_var
        resultat = marge - p["couts_fixes_mensuels"]
        cumul += resultat
        rows.append({
            "mois": m,
            "nouveaux_clients": round(nouveaux, 1),
            "clients_actifs": round(clients, 1),
            "ca_mensuel_eur": round(ca),
            "marge_brute_eur": round(marge),
            "couts_fixes_eur": round(p["couts_fixes_mensuels"]),
            "resultat_mensuel_eur": round(resultat),
            "resultat_cumule_eur": round(cumul),
        })
    return pd.DataFrame(rows)


def point_mort_mois(proj: pd.DataFrame) -> int | None:
    """1er mois ou le resultat mensuel devient positif (>= 0), sinon None."""
    pos = proj[proj["resultat_mensuel_eur"] >= 0]
    return int(pos["mois"].iloc[0]) if not pos.empty else None


MOIS_FR = ["Janvier", "Fevrier", "Mars", "Avril", "Mai", "Juin",
           "Juillet", "Aout", "Septembre", "Octobre", "Novembre", "Decembre"]


def ecrire_feuille(ws, params: dict | None = None, horizon: int = 12) -> None:
    """Modele financier INTERACTIF (formules Excel) inspire d'un modele de
    project-finance : scenario selectionnable, hypotheses editables (jaunes),
    projection (CA -> marge -> EBITDA -> resultat net -> cash-flow -> tresorerie),
    et synthese (point mort, payback, VAN, TRI) + graphe. Tout se recalcule a
    l'edition. (Le calcul Python de projection() reste pour tests / Streamlit.)"""
    from openpyxl.chart import LineChart, Reference
    from openpyxl.styles import Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    import config as cfg

    jaune = PatternFill("solid", fgColor="FFF2CC")
    titre, gras = Font(bold=True, size=14), Font(bold=True)
    EUR, PCT = "# ##0", "0%"

    ws["A1"] = "MODELE FINANCIER PREVISIONNEL CHRUTH"
    ws["A1"].font = titre
    ws["A2"] = ("Choisissez un scenario / modifiez les cellules JAUNES "
                "-> tout se recalcule (hypotheses a valider).")

    # --- Selecteur de scenario + table des scenarios ---
    ws["A4"] = "Scenario selectionne"
    ws["A4"].font = gras
    sel = ws["B4"]
    sel.value = cfg.PREV_SCENARIO_DEFAUT
    sel.fill = jaune
    noms = list(cfg.PREV_SCENARIOS.keys())
    dv = DataValidation(type="list", formula1='"%s"' % ",".join(noms), allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(sel)
    for j, h in enumerate(["Scenario", "Conversion", "Panier annuel", "Couts var %",
                           "Couts fixes/mois"], start=1):
        ws.cell(6, j, h).font = gras
    sc_first = 7
    for k, (nom, (conv, panier, cvar, cfix)) in enumerate(cfg.PREV_SCENARIOS.items()):
        r = sc_first + k
        ws.cell(r, 1, nom)
        ws.cell(r, 2, conv).number_format = PCT
        ws.cell(r, 3, panier).number_format = EUR
        ws.cell(r, 4, cvar).number_format = PCT
        ws.cell(r, 5, cfix).number_format = EUR
    sc_last = sc_first + len(cfg.PREV_SCENARIOS) - 1

    def _lookup(col_letter):
        return (f"=INDEX({col_letter}{sc_first}:{col_letter}{sc_last},"
                f"MATCH($B$4,A{sc_first}:A{sc_last},0))")

    # --- Hypotheses actives (derivees du scenario + editables) ---
    ws["A11"] = "--- Hypotheses actives ---"
    ws["A11"].font = gras
    lignes = [
        ("Prospects travailles / mois", cfg.PREV_CONTACTS_PAR_MOIS, "0", True),
        ("Taux de conversion", _lookup("B"), PCT, False),
        ("Panier moyen annuel / client (EUR)", _lookup("C"), EUR, False),
        ("Couts variables (% du CA)", _lookup("D"), PCT, False),
        ("Couts fixes mensuels (EUR)", _lookup("E"), EUR, False),
        ("Churn mensuel", cfg.PREV_CHURN_MENSUEL, PCT, True),
        ("CAPEX / investissement initial (EUR)", cfg.PREV_CAPEX_INITIAL_EUR, EUR, True),
        ("Taux d'actualisation annuel", cfg.PREV_TAUX_ACTUALISATION, PCT, True),
        ("Taux d'imposition (IS)", cfg.PREV_TAUX_IS, PCT, True),
        ("Amortissement CAPEX (mois)", cfg.PREV_AMORT_MOIS, "0", True),
        ("Annee de l'exercice", cfg.PREV_ANNEE_EXERCICE, "0", True),
    ]
    for i, (label, val, fmt, editable) in enumerate(lignes):
        r = 12 + i
        ws.cell(r, 1, label)
        c = ws.cell(r, 2, val)
        c.number_format = fmt
        if editable:
            c.fill = jaune
    CONTACTS, CONV, PANIER, CVAR, CFIX = "$B$12", "$B$13", "$B$14", "$B$15", "$B$16"
    CHURN, CAPEX, TXACT, TXIS, AMORT = "$B$17", "$B$18", "$B$19", "$B$20", "$B$21"

    # --- Projection (FORMULES) ---
    entete = 25
    headers = ["Mois", "Nouveaux clients", "Clients actifs", "CA mensuel (EUR)",
               "Marge brute (EUR)", "EBITDA (EUR)", "Amortissement (EUR)",
               "Resultat avant impot (EUR)", "Impot (EUR)", "Resultat net (EUR)",
               "Cash-flow (EUR)", "Tresorerie cumulee (EUR)"]
    for j, h in enumerate(headers, start=1):
        ws.cell(entete, j, h).font = gras
    first = entete + 1
    last = first + horizon - 1
    for i in range(horizon):
        rr = first + i
        ws.cell(rr, 1, MOIS_FR[i % 12])
        ws.cell(rr, 2).value = f"={CONTACTS}*{CONV}"
        ws.cell(rr, 3).value = (f"=B{rr}" if i == 0 else f"=C{rr - 1}*(1-{CHURN})+B{rr}")
        ws.cell(rr, 4).value = f"=C{rr}*{PANIER}/12"
        ws.cell(rr, 5).value = f"=D{rr}*(1-{CVAR})"
        ws.cell(rr, 6).value = f"=E{rr}-{CFIX}"                                   # EBITDA
        ws.cell(rr, 7).value = f"=IF(A{rr}<={AMORT},{CAPEX}/{AMORT},0)"           # amortissement
        ws.cell(rr, 8).value = f"=F{rr}-G{rr}"                                    # resultat avant impot
        ws.cell(rr, 9).value = f"=MAX(0,H{rr}*{TXIS})"                            # impot
        ws.cell(rr, 10).value = f"=H{rr}-I{rr}"                                   # resultat net
        ws.cell(rr, 11).value = (f"=J{rr}+G{rr}-{CAPEX}" if i == 0               # cash-flow (CAPEX mois 1)
                                 else f"=J{rr}+G{rr}")
        ws.cell(rr, 12).value = (f"=K{rr}" if i == 0 else f"=L{rr - 1}+K{rr}")    # tresorerie cumulee
        for col in range(4, 13):
            ws.cell(rr, col).number_format = EUR

    # --- Synthese (resultats, colonnes N-O) + graphe ---
    syn = [
        ("Exercice (annee)", "=B22", "0"),
        ("Point mort EBITDA (mois)",
         f'=IFERROR(INDEX(A{first}:A{last},MATCH(TRUE,INDEX(F{first}:F{last}>=0,0),0)),"non atteint")', "General"),
        ("Delai de retour / payback (mois)",
         f'=IFERROR(INDEX(A{first}:A{last},MATCH(TRUE,INDEX(L{first}:L{last}>=0,0),0)),"non atteint")', "General"),
        ("CA total exercice (EUR)", f"=SUM(D{first}:D{last})", EUR),
        ("EBITDA total exercice (EUR)", f"=SUM(F{first}:F{last})", EUR),
        ("Resultat net total exercice (EUR)", f"=SUM(J{first}:J{last})", EUR),
        ("Tresorerie fin d'annee (EUR)", f"=L{last}", EUR),
        ("VAN (EUR)", f"=NPV((1+{TXACT})^(1/12)-1,K{first}:K{last})", EUR),
        ("TRI annuel", f'=IFERROR((1+IRR(K{first}:K{last}))^12-1,"n/a")', PCT),
        ("Clients actifs fin", f"=C{last}", "0"),
    ]
    ws.cell(4, 14, "--- BILAN DE L'EXERCICE ---").font = gras
    for i, (label, formule, fmt) in enumerate(syn):
        r = 5 + i
        ws.cell(r, 14, label).font = gras
        vc = ws.cell(r, 15)
        vc.value = formule
        vc.number_format = fmt

    ch = LineChart()
    ch.title = "CA mensuel & tresorerie cumulee"
    ch.height, ch.width = 8, 18
    for col in (4, 12):
        ch.add_data(Reference(ws, min_col=col, min_row=entete, max_row=last), titles_from_data=True)
    ch.set_categories(Reference(ws, min_col=1, min_row=first, max_row=last))
    ws.add_chart(ch, "N16")

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["N"].width = 30


def hypotheses_synthese(params: dict | None = None) -> pd.DataFrame:
    p = params or parametres()
    proj = projection(p)
    pm = point_mort_mois(proj)
    ca_an1 = proj[proj["mois"] <= 12]["ca_mensuel_eur"].sum()
    ca_an2 = proj[(proj["mois"] > 12) & (proj["mois"] <= 24)]["ca_mensuel_eur"].sum()
    lignes = [
        ("HYPOTHESES (a valider avec CHRUTH)", ""),
        ("Horizon (mois)", p["horizon_mois"]),
        ("Prospects travailles / mois", p["contacts_par_mois"]),
        ("Taux de conversion", f"{p['taux_conversion'] * 100:.0f}%"),
        ("Panier moyen annuel / client (EUR)", p["panier_annuel"]),
        ("Couts variables (% du CA)", f"{p['cout_variable_pct'] * 100:.0f}%"),
        ("Couts fixes mensuels (EUR)", p["couts_fixes_mensuels"]),
        ("Churn mensuel", f"{p['churn_mensuel'] * 100:.0f}%"),
        ("--- RESULTATS PROJETES ---", ""),
        ("Point mort (mois)", pm if pm is not None else "non atteint"),
        ("CA annee 1 (EUR)", round(ca_an1)),
        ("CA annee 2 (EUR)", round(ca_an2)),
        ("Resultat cumule fin d'horizon (EUR)", int(proj["resultat_cumule_eur"].iloc[-1])),
        ("Clients actifs fin d'horizon", round(proj["clients_actifs"].iloc[-1])),
    ]
    return pd.DataFrame(lignes, columns=["Indicateur", "Valeur"])
