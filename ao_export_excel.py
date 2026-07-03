from __future__ import annotations

import os
from pathlib import Path

import pandas as pd

from ao_config import AO_MAX_EXPORT_ROWS, BASE_DIR, IDF_DEPARTEMENTS
from config import departements_des_regions
from ao_db import fetch_logs, fetch_records
from ao_scoring import scoring_rules_table


DISPLAY_COLUMNS = [
    "source",
    "id_ao",
    "objet",
    "acheteur",
    "siret_acheteur",
    "siren_acheteur",
    "nom_contact",
    "prenom_contact",
    "email",
    "telephone",
    "adresse",
    "code_postal",
    "ville",
    "departement",
    "departement_prestation",
    "region",
    "date_publication",
    "date_limite",
    "budget_estime_eur",
    "budget_statut",
    "budget_annuel_eur",
    "budget_annualise",
    "categorie",
    "secteur",
    "type_marche",
    "procedure",
    "nature_avis",
    "descripteur",
    "criteres",
    "url_avis",
    "url_dce",
    "url_profil_acheteur",
    "mots_cles_detectes",
    "nb_infos_disponibles",
    "niveau_confiance",
    "statut_extraction",
    "score_chruth",
    "priorite",
    "score_excel_formule",
    "priorite_excel_formule",
    "qualite_infos_formule",
    "action_excel_formule",
    "ia_statut",
    "ia_action_recommandee",
    "ia_champs_a_completer",
    "ia_prompt_verification",
    "preuve_source",
    "dce_statut",
    "dce_budget",
    "dce_email",
    "dce_tel",
    "dce_contact",
    "dce_resume",
    "raisons_scoring",
    "resume_commercial",
    "proposition_message",
    "script_appel",
    "statut_contact",
    "responsable",
    "date_contact",
    "date_relance",
    "reponse",
    "rdv_obtenu",
    "commentaire_humain",
]


def _ensure_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for column in DISPLAY_COLUMNS:
        if column not in df.columns:
            df[column] = ""
    df = enrich_agent_columns(df)
    return df[DISPLAY_COLUMNS]


def _is_blank(value) -> bool:
    text = str(value or "").strip().lower()
    return text in {"", "nan", "none", "null"}


def _missing_fields(row: pd.Series) -> list[str]:
    fields = {
        "siret_acheteur": "SIRET",
        "email": "email",
        "telephone": "telephone",
        "adresse": "adresse",
        "ville": "ville",
        "budget_estime_eur": "budget",
    }
    return [label for column, label in fields.items() if _is_blank(row.get(column))]


def enrich_agent_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    if df.empty:
        return df

    for idx, row in df.iterrows():
        missing = _missing_fields(row)
        budget_status = str(row.get("budget_statut") or "")
        extraction_status = str(row.get("statut_extraction") or "")

        if extraction_status == "DCE_A_TELECHARGER":
            ia_status = "IA_DCE_REQUISE"
            action = "Telecharger ou ouvrir le DCE, puis extraire contact, budget, perimetre et contraintes."
        elif budget_status == "A_VERIFIER_BUDGET":
            ia_status = "IA_BUDGET_A_VERIFIER"
            action = "Verifier le budget dans l'avis source ou le DCE avant decision commerciale."
        elif missing:
            ia_status = "IA_CONTACT_A_COMPLETER"
            action = "Completer les champs manquants depuis l'avis source, le site acheteur ou le DCE."
        else:
            ia_status = "IA_CONTROLE_SOURCE"
            action = "Verifier rapidement la source avant contact."

        prompt = (
            "Tu es l'agent IA CHRUTH. Verifie cet appel d'offres depuis le lien source. "
            "Extrait uniquement les informations prouvees : acheteur, SIRET, contact, email, telephone, "
            "adresse, ville, budget, date limite, pieces DCE et pertinence pour CHRUTH. "
            "Ne devine pas. Si une information est absente, reponds 'non trouve'. "
            f"ID AO: {row.get('id_ao')}. Objet: {row.get('objet')}. "
            f"Acheteur: {row.get('acheteur')}. Lien: {row.get('url_avis')}."
        )

        df.at[idx, "ia_statut"] = ia_status
        df.at[idx, "ia_action_recommandee"] = action
        df.at[idx, "ia_champs_a_completer"] = ", ".join(missing) if missing else "Aucun champ critique manquant"
        df.at[idx, "ia_prompt_verification"] = prompt
        df.at[idx, "preuve_source"] = "Verifier url_avis et DCE si disponible"

    return df


def _first_department(value) -> str:
    text = str(value or "")
    return text.split(",")[0].strip()


def build_acheteurs(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=["acheteur", "siret_acheteur", "nb_ao", "score_max", "priorites", "villes"])
    work = df.copy()
    work["score_chruth_num"] = pd.to_numeric(work["score_chruth"], errors="coerce").fillna(0)
    return (
        work.groupby(["acheteur", "siret_acheteur"], dropna=False)
        .agg(
            nb_ao=("id_ao", "count"),
            score_max=("score_chruth_num", "max"),
            priorites=("priorite", lambda values: ", ".join(sorted(set(str(v) for v in values if str(v).strip())))),
            villes=("ville", lambda values: ", ".join(list(dict.fromkeys(str(v) for v in values if str(v).strip()))[:8])),
            dernier_ao=("date_publication", "max"),
        )
        .reset_index()
        .sort_values(["score_max", "nb_ao"], ascending=False)
    )


def build_villes(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=["departement", "ville", "nb_ao", "score_moyen", "nb_chaud", "nb_a_verifier"])
    work = df.copy()
    work["dept_prioritaire"] = work["departement_prestation"].where(
        work["departement_prestation"].astype(str).str.strip() != "",
        work["departement"],
    )
    work["dept_prioritaire"] = work["dept_prioritaire"].map(_first_department)
    work["score_chruth_num"] = pd.to_numeric(work["score_chruth"], errors="coerce").fillna(0)
    out = (
        work.groupby(["dept_prioritaire", "ville"], dropna=False)
        .agg(
            nb_ao=("id_ao", "count"),
            score_moyen=("score_chruth_num", "mean"),
            nb_chaud=("priorite", lambda values: int((values == "CHAUD").sum())),
            nb_a_verifier=("priorite", lambda values: int((values == "A_VERIFIER").sum())),
            acheteurs=("acheteur", lambda values: ", ".join(list(dict.fromkeys(str(v) for v in values if str(v).strip()))[:5])),
        )
        .reset_index()
        .rename(columns={"dept_prioritaire": "departement"})
        .sort_values(["nb_chaud", "score_moyen", "nb_ao"], ascending=False)
    )
    out["score_moyen"] = out["score_moyen"].round(1)
    return out


def build_agent_ia() -> pd.DataFrame:
    rows = [
        {
            "cas": "IA_DCE_REQUISE",
            "quand_utiliser": "La ligne manque de contact, budget ou informations critiques.",
            "action_agent": "Lire l'avis source, puis telecharger ou consulter le DCE si necessaire.",
            "sortie_attendue": "Champs complets ou mention 'non trouve' avec preuve.",
        },
        {
            "cas": "IA_BUDGET_A_VERIFIER",
            "quand_utiliser": "Le budget n'est pas explicite dans la ligne Excel.",
            "action_agent": "Chercher montant, valeur estimee, montant maximum ou accord-cadre dans l'avis/DCE.",
            "sortie_attendue": "Budget confirme, budget non trouve, ou budget sous seuil.",
        },
        {
            "cas": "IA_CONTACT_A_COMPLETER",
            "quand_utiliser": "Email, telephone, SIRET, adresse ou ville manquent.",
            "action_agent": "Chercher dans l'avis, le profil acheteur, le site institutionnel et le DCE.",
            "sortie_attendue": "Coordonnees verifiees et source de preuve.",
        },
        {
            "cas": "IA_CONTROLE_SOURCE",
            "quand_utiliser": "La ligne semble complete.",
            "action_agent": "Verifier que l'objet, le delai, le budget et le perimetre correspondent bien a CHRUTH.",
            "sortie_attendue": "Decision : contacter, surveiller, non pertinent.",
        },
    ]
    return pd.DataFrame(rows)


def vba_module_text() -> str:
    return r'''Attribute VB_Name = "AO_CHRUTH"
Option Explicit

Sub Update_AO_CHRUTH()
    Dim projet As String
    Dim region As String
    Dim cmd As String
    projet = ThisWorkbook.Path & "\.."
    region = ""
    On Error Resume Next
    region = Trim(CStr(ThisWorkbook.Sheets("Parametres").Range("B1").Value))
    On Error GoTo 0
    cmd = "cmd.exe /c cd /d """ & projet & """ && python outils\refresh_runner.py ao"
    If Len(region) > 0 Then cmd = cmd & " """ & region & """"
    Shell cmd, vbNormalFocus
    ThisWorkbook.Close SaveChanges:=False
End Sub

Sub Aller_AO_Region()
    On Error Resume Next
    Sheets("AO_Region").Activate
    On Error GoTo 0
End Sub

Sub Aller_AO_Chauds()
    Sheets("AO_CHAUDS").Activate
End Sub

Sub Aller_CRM()
    Sheets("CRM_Suivi").Activate
End Sub

Sub Filtrer_AO_Chauds()
    Sheets("AO_Tous").Activate
    If ActiveSheet.AutoFilterMode Then ActiveSheet.AutoFilterMode = False
    Range("A1").CurrentRegion.AutoFilter Field:=33, Criteria1:="CHAUD"
End Sub

Sub Basculer_Collecte_Donnees()
    ' Active / desactive la collecte reseau CHRUTH.
    ' OFF = les boutons retraitent/exportent les donnees locales sans appeler BOAMP/DCE.
    Dim ws As Worksheet
    Dim etat As String
    Dim nouveau As String
    Dim projet As String
    On Error Resume Next
    Set ws = ThisWorkbook.Sheets("Parametres")
    On Error GoTo 0
    If ws Is Nothing Then
        MsgBox "Onglet Parametres introuvable.", vbExclamation
        Exit Sub
    End If
    etat = UCase(Trim(CStr(ws.Range("B3").Value)))
    If etat = "OFF" Then
        nouveau = "ON"
    Else
        nouveau = "OFF"
    End If
    ws.Range("B3").Value = nouveau
    If nouveau = "ON" Then
        ws.Range("B3").Interior.Color = RGB(198, 239, 206)
    Else
        ws.Range("B3").Interior.Color = RGB(255, 199, 206)
    End If
    projet = ThisWorkbook.Path & "\.."
    Shell "cmd.exe /c cd /d """ & projet & """ && python outils\set_collecte.py " & nouveau, vbHide
    ThisWorkbook.Save
    MsgBox "Collecte de donnees : " & nouveau, vbInformation
End Sub

Sub Basculer_Notifications()
    ' Active / desactive les alertes email CHRUTH (coupe les emails seulement ;
    ' la collecte et la mise a jour du tableur continuent). L'etat est persiste
    ' cote Python via outils\set_notifications.py pour survivre aux refresh.
    Dim ws As Worksheet
    Dim etat As String
    Dim nouveau As String
    Dim projet As String
    On Error Resume Next
    Set ws = ThisWorkbook.Sheets("Parametres")
    On Error GoTo 0
    If ws Is Nothing Then
        MsgBox "Onglet Parametres introuvable.", vbExclamation
        Exit Sub
    End If
    etat = UCase(Trim(CStr(ws.Range("B2").Value)))
    If etat = "OFF" Then
        nouveau = "ON"
    Else
        nouveau = "OFF"
    End If
    ws.Range("B2").Value = nouveau
    If nouveau = "ON" Then
        ws.Range("B2").Interior.Color = RGB(198, 239, 206)
    Else
        ws.Range("B2").Interior.Color = RGB(255, 199, 206)
    End If
    projet = ThisWorkbook.Path & "\.."
    Shell "cmd.exe /c cd /d """ & projet & """ && python outils\set_notifications.py " & nouveau, vbHide
    ThisWorkbook.Save
    MsgBox "Notifications email : " & nouveau, vbInformation
End Sub

Sub Enregistrer_Destinataires()
    ' Enregistre la liste des destinataires saisie dans l'onglet Parametres
    ' (colonne B a partir de la ligne 5) pour les prochaines alertes email.
    Dim projet As String
    Dim sh As Object
    Dim rc As Long
    ThisWorkbook.Save
    projet = ThisWorkbook.Path & "\.."
    Set sh = CreateObject("WScript.Shell")
    rc = sh.Run("cmd /c cd /d """ & projet & """ && python outils\sync_destinataires.py", 0, True)
    If rc = 0 Then
        MsgBox "Destinataires enregistres. Ils recevront les prochaines alertes.", vbInformation
    Else
        MsgBox "Aucune adresse valide trouvee (colonne B a partir de la ligne 5).", vbExclamation
    End If
End Sub

Sub Generer_Message_AO()
    ' Genere le message IA (email + script) pour l'AO de la ligne selectionnee.
    ' Lit id_ao sur la feuille active, appelle Python (Ollama), affiche le resultat
    ' dans l'onglet Message_IA (a copier/coller).
    Dim idCol As Long, lastCol As Long, c As Long, idAO As String
    Dim projet As String, sh As Object, rc As Long, fichier As String
    Dim ws As Worksheet, stream As Object, contenu As String, lignes() As String, i As Long
    On Error GoTo Erreur

    lastCol = Cells(1, Columns.Count).End(xlToLeft).Column
    For c = 1 To lastCol
        If LCase(Trim(CStr(Cells(1, c).Value))) = "id_ao" Then idCol = c: Exit For
    Next c
    If idCol = 0 Then
        MsgBox "Place-toi sur l'onglet AO_Nettoyage_IDF (colonne id_ao requise).", vbExclamation
        Exit Sub
    End If
    If ActiveCell.Row < 2 Then
        MsgBox "Selectionne d'abord une ligne d'appel d'offres.", vbInformation
        Exit Sub
    End If
    idAO = Trim(CStr(Cells(ActiveCell.Row, idCol).Value))
    If idAO = "" Then
        MsgBox "Pas d'identifiant sur cette ligne.", vbInformation
        Exit Sub
    End If

    projet = ThisWorkbook.Path & "\.."
    Application.StatusBar = "Generation du message IA en cours (Ollama)..."
    Set sh = CreateObject("WScript.Shell")
    rc = sh.Run("cmd /c cd /d """ & projet & """ && python outils\generer_message_ao.py """ & idAO & """", 0, True)
    Application.StatusBar = False

    fichier = projet & "\output\_message_ao.txt"
    If Dir(fichier) = "" Then
        MsgBox "Generation echouee (fichier non produit). Verifie Python/Ollama.", vbExclamation
        Exit Sub
    End If

    Set stream = CreateObject("ADODB.Stream")
    stream.Charset = "utf-8"
    stream.Open
    stream.LoadFromFile fichier
    contenu = stream.ReadText
    stream.Close

    On Error Resume Next
    Set ws = ThisWorkbook.Sheets("Message_IA")
    On Error GoTo Erreur
    If ws Is Nothing Then
        Set ws = ThisWorkbook.Sheets.Add(After:=ThisWorkbook.Sheets(ThisWorkbook.Sheets.Count))
        ws.Name = "Message_IA"
    End If
    ws.Cells.Clear
    ws.Columns(1).NumberFormat = "@"   ' texte : sinon "===..." est pris pour une formule (erreur 1004)
    lignes = Split(Replace(contenu, vbCrLf, vbLf), vbLf)
    For i = 0 To UBound(lignes)
        ws.Cells(i + 1, 1).Value = lignes(i)
    Next i
    ws.Columns(1).ColumnWidth = 110
    ws.Activate
    ws.Range("A1").Select
    MsgBox "Message genere dans l'onglet Message_IA (copier/coller depuis la colonne A).", vbInformation
    Exit Sub
Erreur:
    Application.StatusBar = False
    MsgBox "Erreur generation message : " & Err.Number & " - " & Err.Description, vbExclamation
End Sub

Sub Ouvrir_Lien_AO()
    Dim urlCol As Long
    Dim lastCol As Long
    Dim c As Long
    lastCol = Cells(1, Columns.Count).End(xlToLeft).Column
    For c = 1 To lastCol
        If Cells(1, c).Value = "url_avis" Then
            urlCol = c
            Exit For
        End If
    Next c
    If urlCol = 0 Then
        MsgBox "Colonne url_avis introuvable.", vbExclamation
        Exit Sub
    End If
    If ActiveCell.Row < 2 Then
        MsgBox "Selectionne une ligne AO.", vbInformation
        Exit Sub
    End If
    If Cells(ActiveCell.Row, urlCol).Value <> "" Then
        ThisWorkbook.FollowHyperlink Cells(ActiveCell.Row, urlCol).Value
    Else
        MsgBox "Aucun lien source sur cette ligne.", vbInformation
    End If
End Sub
'''


def build_vba_sheet() -> pd.DataFrame:
    lines = vba_module_text().splitlines()
    rows = [
        {"section": "1. Utilisation", "contenu": "Enregistrer une copie du fichier en .xlsm, puis importer AO_CHRUTH_VBA.bas dans l'editeur VBA."},
        {"section": "2. Bouton", "contenu": "Creer une forme ou un bouton Excel et affecter la macro Update_AO_CHRUTH."},
        {"section": "2b. Bouton notifications", "contenu": "Creer un 2e bouton (ex. dans l'onglet Parametres) et affecter la macro Basculer_Notifications : il bascule Parametres!B2 entre ON/OFF (vert/rouge) et coupe/reactive les emails d'alerte. La cellule B2 affiche l'etat courant."},
        {"section": "2c. Bouton collecte", "contenu": "Creer un bouton dans Parametres et affecter la macro Basculer_Collecte_Donnees : il bascule Parametres!B3 entre ON/OFF et coupe/reactive la collecte reseau."},
        {"section": "3. Securite", "contenu": "Excel peut bloquer les macros. Activer seulement si le fichier reste local et controle."},
        {"section": "4. Code VBA", "contenu": "Le code complet est aussi exporte dans AO_CHRUTH_VBA.bas."},
    ]
    rows.extend({"section": "VBA", "contenu": line} for line in lines)
    return pd.DataFrame(rows)


def build_dce_a_recuperer(df: pd.DataFrame) -> pd.DataFrame:
    cols = ["id_ao", "objet", "acheteur", "url_dce", "url_profil_acheteur", "dce_statut", "fichier_attendu"]
    if df is None or df.empty:
        return pd.DataFrame(columns=cols)

    def _blank(v):
        return str(v or "").strip().lower() in {"", "nan", "none", "null"}

    work = df.copy()
    mask = work.apply(lambda r: _blank(r.get("budget_estime_eur")) or _blank(r.get("email")) or _blank(r.get("telephone")), axis=1)
    work = work[mask].copy()
    if work.empty:
        return pd.DataFrame(columns=cols)
    work["score_chruth_num"] = pd.to_numeric(work.get("score_chruth"), errors="coerce").fillna(0)
    work = work.sort_values("score_chruth_num", ascending=False)
    work["fichier_attendu"] = work["id_ao"].astype(str) + ".pdf"
    statut = work.get("dce_statut")
    if statut is None:
        work["dce_statut"] = ""
    work["dce_statut"] = work.apply(
        lambda r: r["dce_statut"] if not _blank(r.get("dce_statut")) else ("LIEN_SEUL" if not _blank(r.get("url_dce")) else "AUCUN_LIEN"),
        axis=1,
    )
    for c in cols:
        if c not in work.columns:
            work[c] = ""
    return work[cols].reset_index(drop=True)


ESSENTIELS_COLS = [
    ("objet", "Objet de l'AO"),
    ("acheteur", "Acheteur"),
    ("secteur", "Secteur"),
    ("categorie", "Catégorie"),
    ("ville", "Ville"),
    ("departement_prestation", "Département"),
    ("date_publication", "Date de publication"),
    ("date_limite", "Date limite de réponse"),
    ("budget_annuel_eur", "Budget estimé (€)"),
    ("url_dce", "Lien DCE"),
    ("url_avis", "Lien avis BOAMP"),
    ("priorite", "Priorité"),
    ("score_chruth", "Score"),
    ("id_ao", "id_ao"),  # cle technique : lue par le bouton "Generer le message IA"
]


def build_essentiels(df: pd.DataFrame) -> pd.DataFrame:
    """Vue epuree demandee par le client (rien n'est supprime : detail complet dans AO_Tous)."""
    cols = [src for src, _ in ESSENTIELS_COLS]
    work = df.copy()
    for src in cols:
        if src not in work.columns:
            work[src] = ""
    # Budget affiche : annualise si dispo, sinon total.
    work["budget_annuel_eur"] = work.apply(
        lambda r: r["budget_annuel_eur"] if str(r.get("budget_annuel_eur") or "").strip() not in ("", "nan") else r.get("budget_estime_eur", ""),
        axis=1,
    )
    out = work[cols].copy()
    out.columns = [label for _, label in ESSENTIELS_COLS]
    # Mission 3 : brouillons IA a l'acheteur (ajoutes si la generation a tourne).
    if "brouillon_email_ia" in work.columns:
        out["Brouillon email (IA)"] = work["brouillon_email_ia"].values
        out["Brouillon script (IA)"] = work["brouillon_script_ia"].values
    return out


def _write_df_to_sheet(ws, df) -> None:
    from openpyxl.utils.dataframe import dataframe_to_rows
    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(["" if v is None else v for v in r])


SHEET_ORDER = [
    "Pilotage", "Top20_Ouverts", "AO_Nettoyage_IDF", "AO_Tous", "AO_Ile_de_France", "AO_Region", "AO_CHAUDS",
    "AO_A_VERIFIER", "DCE_A_RECUPERER", "CRM_Suivi", "Agent_IA", "Scoring", "Acheteurs",
    "Villes_Departements", "VBA_Bouton", "Logs_Update",
]


def export_excel(path: Path = None, region_names: str = "") -> Path:
    from datetime import date
    from openpyxl import load_workbook
    from ao_config import AO_TEMPLATE_XLSM, AO_OUTPUT_XLSM
    from ao_pilotage import compute_kpis, render_pilotage
    from ao_semaine import add_semaine_iso, build_top20_ouverts

    path = path or AO_OUTPUT_XLSM
    path.parent.mkdir(parents=True, exist_ok=True)

    region_departements = None
    region_label = ""
    if region_names.strip():
        noms = [r.strip() for r in region_names.split(",") if r.strip()]
        region_departements = departements_des_regions(noms)
        region_label = ", ".join(noms)

    raw = fetch_records()
    if not raw.empty:
        raw["score_chruth_num"] = pd.to_numeric(raw["score_chruth"], errors="coerce").fillna(0)
        raw = raw.sort_values(["score_chruth_num", "date_publication"], ascending=False).drop(columns=["score_chruth_num"])
    # df_full = toute la base triee (pour la vue region) ; df = top-100 (feuilles globales).
    df_full = _ensure_columns(raw)
    df_full = add_semaine_iso(df_full)
    df = df_full.head(AO_MAX_EXPORT_ROWS)

    today = date.today()
    dept = df["departement_prestation"].where(df["departement_prestation"].astype(str).str.strip() != "", df["departement"])
    idf_mask = dept.map(_first_department).isin(IDF_DEPARTEMENTS)
    # AO_Region : construit sur TOUTE la base filtree region (pas seulement le top-100
    # global), puis cap au top-100 de la region -> vraie focalisation regionale.
    if region_departements:
        dept_full = df_full["departement_prestation"].where(
            df_full["departement_prestation"].astype(str).str.strip() != "", df_full["departement"])
        ao_region = df_full[dept_full.map(_first_department).isin(region_departements)].head(AO_MAX_EXPORT_ROWS)
    else:
        ao_region = df  # pas de region selectionnee -> vue identique au releve global
    hot = df[df["priorite"] == "CHAUD"]
    to_check = df[
        (df["priorite"] == "A_VERIFIER")
        | (df["budget_statut"] == "A_VERIFIER_BUDGET")
        | (df["statut_extraction"] == "DCE_A_TELECHARGER")
    ]
    crm = df[df["priorite"].isin(["CHAUD", "A_VERIFIER", "TIEDE"])].copy()
    top20 = build_top20_ouverts(df, today=today)
    kpis = compute_kpis(df, today=today)

    # Mission 3 (volet AO) : brouillons IA a l'acheteur pour les AO IDF CHAUD/TIEDE.
    idf_df = df[idf_mask]
    if os.environ.get("CHRUTH_GENERER_MESSAGES", "").strip() in ("1", "true", "True"):
        try:
            import ao_messages
            idf_df = ao_messages.generer_pour_ao_df(idf_df)
        except Exception as exc:  # noqa: BLE001
            print(f"[MESSAGES AO] generation ignoree : {exc}")

    data_sheets = {
        "AO_Nettoyage_IDF": build_essentiels(idf_df),
        "Top20_Ouverts": top20,
        "AO_Tous": df,
        "AO_Ile_de_France": df[idf_mask],
        "AO_Region": ao_region,
        "AO_CHAUDS": hot,
        "AO_A_VERIFIER": to_check,
        "DCE_A_RECUPERER": build_dce_a_recuperer(df),
        "CRM_Suivi": crm,
        "Agent_IA": build_agent_ia(),
        "Scoring": scoring_rules_table(),
        "Acheteurs": build_acheteurs(df),
        "Villes_Departements": build_villes(df),
        "VBA_Bouton": build_vba_sheet(),
        "Logs_Update": fetch_logs(),
    }

    wb = load_workbook(AO_TEMPLATE_XLSM, keep_vba=True)
    pil = wb["Pilotage"] if "Pilotage" in wb.sheetnames else wb.create_sheet("Pilotage", 0)
    for row in pil.iter_rows():
        for cell in row:
            cell.value = None
    render_pilotage(pil, kpis, top20)

    for name, frame in data_sheets.items():
        ws = wb[name] if name in wb.sheetnames else wb.create_sheet(name)
        _write_df_to_sheet(ws, frame)

    # Refleter la region collectee dans la cellule de selection (onglet Parametres).
    if region_label and "Parametres" in wb.sheetnames:
        wb["Parametres"]["B1"] = region_label

    # Refleter les etats ON/OFF dans Parametres.
    # La source de verite reste les fichiers drapeau ; on ne fait qu'afficher l'etat.
    if "Parametres" in wb.sheetnames:
        from openpyxl.styles import PatternFill

        from ao_config import collecte_active, notifications_actives
        notif_actif = notifications_actives()
        collecte_actif = collecte_active()
        param = wb["Parametres"]
        param["A2"] = "Notifications email"
        param["B2"] = "ON" if notif_actif else "OFF"
        param["B2"].fill = PatternFill("solid", fgColor="C6EFCE" if notif_actif else "FFC7CE")
        param["A3"] = "Collecte donnees"
        param["B3"] = "ON" if collecte_actif else "OFF"
        param["B3"].fill = PatternFill("solid", fgColor="C6EFCE" if collecte_actif else "FFC7CE")
        _remplir_destinataires(param)

    improve_format_wb(wb)

    ordered = [n for n in SHEET_ORDER if n in wb.sheetnames] + [n for n in wb.sheetnames if n not in SHEET_ORDER]
    wb._sheets.sort(key=lambda s: ordered.index(s.title))

    write_vba_module()
    wb.save(path)
    return path


def _remplir_destinataires(param) -> None:
    """Repeuple la colonne destinataires de l'onglet Parametres depuis destinataires.txt
    (la source persistante), pour que les adresses survivent a la regeneration."""
    from ao_config import (
        ALERTE_DESTINATAIRES_FILE,
        PARAM_DEST_COL,
        PARAM_DEST_FIRST_ROW,
        PARAM_DEST_MAX,
    )

    f = ALERTE_DESTINATAIRES_FILE
    emails: list[str] = []
    if f.exists():
        for line in f.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line and not line.startswith("#") and "@" in line:
                emails.append(line)
    param[f"A{PARAM_DEST_FIRST_ROW - 1}"] = "Destinataires des alertes (une adresse par ligne)"
    for i in range(PARAM_DEST_MAX):
        cell = f"{PARAM_DEST_COL}{PARAM_DEST_FIRST_ROW + i}"
        param[cell] = emails[i] if i < len(emails) else None


def improve_format_wb(wb) -> None:
    from ao_style import style_data_sheet
    for ws in wb.worksheets:
        if ws.title in ("Pilotage", "Parametres"):
            continue
        headers = {c.value: i for i, c in enumerate(ws[1], start=1)} if ws.max_row >= 1 else {}
        if ws.title == "AO_Tous" and ws.max_row > 1 and headers:
            add_excel_formulas(ws, headers)
        freeze = "objet" if "objet" in headers else None
        style_data_sheet(ws, freeze_until=freeze)


def write_vba_module() -> Path:
    target = BASE_DIR / "AO_CHRUTH_VBA.bas"
    target.write_text(vba_module_text(), encoding="utf-8")
    return target


def add_excel_formulas(ws, headers: dict[str, int]) -> None:
    from openpyxl.utils import get_column_letter

    required = ["objet", "departement", "departement_prestation", "budget_estime_eur", "budget_statut", "email", "telephone", "siret_acheteur", "adresse", "ville", "url_avis", "date_limite", "score_excel_formule", "priorite_excel_formule", "qualite_infos_formule", "action_excel_formule", "statut_extraction"]
    if not all(name in headers for name in required):
        return

    obj = get_column_letter(headers["objet"])
    dep = get_column_letter(headers["departement"])
    dep_prest = get_column_letter(headers["departement_prestation"])
    budget = get_column_letter(headers["budget_estime_eur"])
    budget_status = get_column_letter(headers["budget_statut"])
    email = get_column_letter(headers["email"])
    phone = get_column_letter(headers["telephone"])
    siret = get_column_letter(headers["siret_acheteur"])
    address = get_column_letter(headers["adresse"])
    city = get_column_letter(headers["ville"])
    url = get_column_letter(headers["url_avis"])
    deadline = get_column_letter(headers["date_limite"])
    score_col = get_column_letter(headers["score_excel_formule"])
    priority_col = get_column_letter(headers["priorite_excel_formule"])
    qualite_col = get_column_letter(headers["qualite_infos_formule"])
    action_col = get_column_letter(headers["action_excel_formule"])
    extraction_col = get_column_letter(headers["statut_extraction"])

    idf_checks = ",".join([f'--({dep}{{row}}="{d}")+--({dep_prest}{{row}}="{d}")' for d in IDF_DEPARTEMENTS])
    for row in range(2, ws.max_row + 1):
        semantic = f'IF(OR(ISNUMBER(SEARCH("nettoyage",{obj}{row})),ISNUMBER(SEARCH("proprete",{obj}{row})),ISNUMBER(SEARCH("entretien",{obj}{row})),ISNUMBER(SEARCH("hygiene",{obj}{row}))),35,0)'
        geo = f"IF(SUM({idf_checks.format(row=row)})>0,25,4)"
        budget_formula = (
            f'IF({budget}{row}="",10,'
            f'IF({budget}{row}<=50000,25,'
            f'IF({budget}{row}<=100000,20,'
            f'IF({budget}{row}<=200000,5,'
            f'IF({budget}{row}<=500000,-10,-20)))))'
        )
        info = f'IF({email}{row}<>"",4,0)+IF({phone}{row}<>"",4,0)+IF({siret}{row}<>"",3,0)+IF({address}{row}<>"",2,0)+IF({city}{row}<>"",1,0)+IF({url}{row}<>"",1,0)'
        timing = f'IF({deadline}{row}="",0,IF({deadline}{row}<TODAY(),-30,IF({deadline}{row}<TODAY()+5,-15,IF({deadline}{row}<=TODAY()+15,5,10))))'
        ws[f"{score_col}{row}"] = f"=MIN(100,MAX(0,{semantic}+{geo}+{budget_formula}+{info}+{timing}))"
        ws[f"{priority_col}{row}"] = f'=IF({score_col}{row}>=65,"CHAUD",IF({score_col}{row}>=40,"TIEDE","FROID"))'

        present = (
            f'IF({email}{row}<>"",1,0)+IF({phone}{row}<>"",1,0)'
            f'+IF({siret}{row}<>"",1,0)+IF({address}{row}<>"",1,0)'
        )
        ws[f"{qualite_col}{row}"] = (
            f'=IF(({present})>=4,"Complet",IF(({present})>=2,"Partiel","Vide"))'
        )
        ws[f"{action_col}{row}"] = (
            f'=IF({extraction_col}{row}="DCE_A_TELECHARGER","Telecharger DCE",'
            f'IF({budget_status}{row}="A_VERIFIER_BUDGET","Verifier budget",'
            f'IF({priority_col}{row}="CHAUD","Contacter maintenant",'
            f'IF({priority_col}{row}="TIEDE","Surveiller / relancer","Basse priorite"))))'
        )


def main() -> int:
    import argparse
    parser = argparse.ArgumentParser(description="Export Excel AO CHRUTH.")
    parser.add_argument("--regions", default="",
                        help="Region(s) a mettre en avant dans l'onglet AO_Region, ex: \"Île-de-France\".")
    args = parser.parse_args()
    path = export_excel(region_names=args.regions)
    print(f"Excel AO exporte : {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
