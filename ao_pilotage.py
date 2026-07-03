from __future__ import annotations

from datetime import date

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ao_config import IDF_DEPARTEMENTS
from ao_style import DATE_FMT, EURO_FMT, coerce_date, coerce_numeric

NOM_PROJET = "Prospection Appels d'Offres CHRUTH"

NAVY = "1F3864"
NAVY_LIGHT = "2E5496"
WHITE = "FFFFFF"
CARD_FILL = "D9E1F2"

_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_KPI_ORDER = [
    ("Date MAJ", "date_maj"),
    ("Check qualite", "check_qualite"),
    ("AO trouves", "nb_ao"),
    ("AO chauds", "nb_chauds"),
    ("AO IDF", "nb_idf"),
    ("Budget a verifier", "budget_a_verifier"),
]


def _first_department(value) -> str:
    return str(value or "").split(",")[0].strip()


def compute_kpis(df: pd.DataFrame, today: date | None = None) -> dict:
    today = today or date.today()
    if df is None or df.empty:
        return {
            "nom_projet": NOM_PROJET,
            "date_maj": today.strftime("%d/%m/%Y"),
            "check_qualite": "OK",
            "nb_ao": 0,
            "nb_chauds": 0,
            "nb_idf": 0,
            "budget_a_verifier": 0,
        }

    dept = df["departement_prestation"].where(
        df["departement_prestation"].astype(str).str.strip() != "", df["departement"]
    )
    idf = dept.map(_first_department).isin(IDF_DEPARTEMENTS)
    budget_a_verifier = int((df["budget_statut"] == "A_VERIFIER_BUDGET").sum())
    dce = int((df["statut_extraction"] == "DCE_A_TELECHARGER").sum())
    alertes = budget_a_verifier + dce

    return {
        "nom_projet": NOM_PROJET,
        "date_maj": today.strftime("%d/%m/%Y"),
        "check_qualite": "OK" if alertes == 0 else f"{alertes} alertes",
        "nb_ao": int(len(df)),
        "nb_chauds": int((df["priorite"] == "CHAUD").sum()),
        "nb_idf": int(idf.sum()),
        "budget_a_verifier": budget_a_verifier,
    }


def render_pilotage(ws: Worksheet, kpis: dict, top20) -> None:
    # 1) Bandeau titre
    ws.merge_cells("A1:J2")
    title = ws["A1"]
    title.value = f"PILOTAGE — {kpis.get('nom_projet', '')}"
    title.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    title.font = Font(bold=True, size=16, color=WHITE)
    title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22

    # 2) Cartes KPI (ligne 4 label, ligne 5 valeur), 2 colonnes par carte
    start_row = 4
    for i, (label, key) in enumerate(_KPI_ORDER):
        c1 = 1 + i * 2
        lab = ws.cell(row=start_row, column=c1, value=label)
        ws.merge_cells(start_row=start_row, start_column=c1, end_row=start_row, end_column=c1 + 1)
        lab.fill = PatternFill(start_color=NAVY_LIGHT, end_color=NAVY_LIGHT, fill_type="solid")
        lab.font = Font(bold=True, color=WHITE, size=9)
        lab.alignment = Alignment(horizontal="center", vertical="center")
        val = ws.cell(row=start_row + 1, column=c1, value=kpis.get(key, ""))
        ws.merge_cells(start_row=start_row + 1, start_column=c1, end_row=start_row + 1, end_column=c1 + 1)
        val.fill = PatternFill(start_color=CARD_FILL, end_color=CARD_FILL, fill_type="solid")
        val.font = Font(bold=True, size=14, color=NAVY)
        val.alignment = Alignment(horizontal="center", vertical="center")
        val.border = _BORDER

    # 3) Titre bloc Top 20
    block_row = start_row + 3
    bc = ws.cell(row=block_row, column=1, value="Top 20 AO ouverts (date limite a venir, par score)")
    bc.font = Font(bold=True, size=12, color=NAVY)

    # 4) Tableau Top 20
    header_row = block_row + 1
    headers = list(top20.columns)
    for j, name in enumerate(headers, start=1):
        h = ws.cell(row=header_row, column=j, value=name)
        h.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
        h.font = Font(bold=True, color=WHITE, size=10)
        h.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r, (_, row) in enumerate(top20.iterrows(), start=header_row + 1):
        for j, name in enumerate(headers, start=1):
            value = row[name]
            cell = ws.cell(row=r, column=j)
            if name == "budget_estime_eur" and (num := coerce_numeric(value)) is not None:
                cell.value = num
                cell.number_format = EURO_FMT
            elif name == "date_limite" and (parsed := coerce_date(value)) is not None:
                cell.value = parsed
                cell.number_format = DATE_FMT
            else:
                cell.value = value
            cell.border = _BORDER
            if name == "url_avis" and value:
                cell.hyperlink = str(value)
                cell.style = "Hyperlink"
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate

    # 5) Largeurs colonnes (indexees par nom : robuste a la presence d'id_ao)
    width_map = {
        "rang": 6, "id_ao": 14, "score_chruth": 8, "priorite": 12, "objet": 45,
        "acheteur": 28, "ville": 16, "departement": 8, "budget_estime_eur": 14,
        "date_limite": 14, "url_avis": 30,
    }
    for idx, name in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width_map.get(name, 14)
