"""Ajoute la liste deroulante des regions a l'onglet 'Parametres' du gabarit AO.

A lancer APRES creer_template_xlsm.ps1 (qui cree l'onglet Parametres + la VBA).
On passe par openpyxl (UTF-8) pour ecrire les noms de regions accentues
proprement, et on reference une plage (D2:D19) pour contourner la limite des
255 caracteres d'une liste inline de validation de donnees.
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

BASE = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE))

from config import REGIONS_DEPARTEMENTS  # noqa: E402

TEMPLATE = BASE / "assets" / "AO_CHRUTH_TEMPLATE.xlsm"
REGION_DEFAUT = "Île-de-France"


def main() -> int:
    if not TEMPLATE.exists():
        raise SystemExit(f"Gabarit introuvable : {TEMPLATE}. Lance d'abord creer_template_xlsm.ps1.")

    wb = load_workbook(TEMPLATE, keep_vba=True)
    if "Parametres" not in wb.sheetnames:
        ws = wb.create_sheet("Parametres")
    else:
        ws = wb["Parametres"]

    ws["A1"] = "Region a mettre a jour (AO)"
    ws["B1"] = REGION_DEFAUT

    # Liste de reference des regions en colonne D (D1 = entete, D2..Dn = valeurs).
    regions = list(REGIONS_DEPARTEMENTS.keys())
    ws["D1"] = "Regions_disponibles"
    for i, nom in enumerate(regions, start=2):
        ws.cell(row=i, column=4, value=nom)
    derniere = len(regions) + 1  # ligne de la derniere region

    # Validation de donnees : liste pointant vers la plage (contourne limite 255 car.).
    dv = DataValidation(
        type="list",
        formula1=f"=Parametres!$D$2:$D${derniere}",
        allow_blank=True,
        showDropDown=False,  # False => la fleche du menu deroulant EST affichee (quirk openpyxl)
    )
    dv.prompt = "Choisis une region puis clique sur 'Mettre a jour les AO'."
    dv.promptTitle = "Region AO"
    ws.add_data_validation(dv)
    dv.add(ws["B1"])

    # Largeurs lisibles.
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["D"].width = 26

    wb.save(TEMPLATE)
    print(f"Liste deroulante ajoutee ({len(regions)} regions) -> {TEMPLATE}")
    print(f"Cellule B1 = '{REGION_DEFAUT}' (defaut). Plage source : D2:D{derniere}.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
