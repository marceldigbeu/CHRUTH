"""Post-traitement du modele financier CHRUTH."""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

NAVY = "17324D"
TEAL = "0F766E"
LIGHT_BLUE = "D9EAF7"
LIGHT_GRAY = "F3F6F8"
INPUT = "FFF2CC"
WHITE = "FFFFFF"
DARK = "1F2933"


def _header(cells, fill=TEAL):
    for cell in cells:
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _title(cell):
    cell.font = Font(bold=True, size=16, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY)


def _clear(ws):
    ws.delete_rows(1, ws.max_row)
    ws._charts = []
    for table_name in list(ws.tables.keys()):
        del ws.tables[table_name]


def ameliorer(path: str | Path) -> Path:
    path = Path(path)
    wb = load_workbook(path)
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except Exception:
        pass

    pre = wb["Previsionnel_Global"]
    marches = wb["Marches"]
    synth = wb["Synthese_Rentabilite"]
    hyp = wb["Hypotheses_Marche"]

    pre["A2"] = "Choisir un scenario et modifier les cellules jaunes. Les indicateurs, la tresorerie et le point mort se recalculent a l ouverture."
    for row in range(26, 38):
        month_index = row - 25
        pre[f"G{row}"] = f"=IF({month_index}<=$B$21,$B$18/$B$21,0)"
        pre[f"K{row}"] = "=J26+G26-$B$18" if row == 26 else f"=J{row}+G{row}"
        pre[f"L{row}"] = "=K26" if row == 26 else f"=L{row-1}+K{row}"
    pre.data_validations.dataValidation = []
    dv = DataValidation(type="list", formula1="=$A$7:$A$9", allow_blank=False)
    pre.add_data_validation(dv)
    dv.add(pre["B4"])
    pre.freeze_panes = "A25"

    hyp["A9"] = "Prix minimum = cout total / (1 - marge cible)"
    hyp["B9"] = "=1/(1-B7)"
    hyp["A10"] = "Alerte marge basse si marge < cible"
    hyp["B10"] = "=B7"

    for table_name in list(marches.tables.keys()):
        del marches.tables[table_name]
    headers = [
        "Nom du marche / client", "Jours/sem", "Heures/jour", "Effectif", "Type", "Produits par",
        "Cout produits/mois", "Prix du marche/mois", "Amortissement/mois", "Heures/mois",
        "Cout main d'oeuvre", "Cout produits", "Cout total", "Prix minimum cible", "Marge EUR",
        "Marge %", "Statut marge", "Priorite business", "Ecart vs prix minimum", "Decision / action",
    ]
    for col, label in enumerate(headers, 1):
        marches.cell(3, col, label)
    _header(marches[3][0:len(headers)])
    max_row = 25
    for row in range(4, max_row + 1):
        if row > 15:
            for col in range(1, len(headers) + 1):
                marches.cell(row, col).value = None
        if marches[f"E{row}"].value in (None, ""):
            marches[f"E{row}"] = "Jour"
        if marches[f"F{row}"].value in (None, ""):
            marches[f"F{row}"] = "Nous"
        marches[f"J{row}"] = f'=IF($A{row}="","",$B{row}*$C{row}*$D{row}*Hypotheses_Marche!$B$5)'
        marches[f"K{row}"] = f'=IF($A{row}="","",$J{row}*IF($E{row}="Nuit",Hypotheses_Marche!$B$4,Hypotheses_Marche!$B$3))'
        marches[f"L{row}"] = f'=IF($A{row}="","",IF(AND($F{row}="Nous",$H{row}<>""),IF($G{row}="",Hypotheses_Marche!$B$6,$G{row}),0))'
        marches[f"M{row}"] = f'=IF($A{row}="","",$K{row}+$L{row}+$I{row})'
        marches[f"N{row}"] = f'=IF($A{row}="","",$M{row}/(1-Hypotheses_Marche!$B$7))'
        marches[f"O{row}"] = f'=IF($A{row}="","",$H{row}-$M{row})'
        marches[f"P{row}"] = f'=IF(OR($A{row}="",$H{row}=0),"",$O{row}/$H{row})'
        marches[f"Q{row}"] = f'=IF($A{row}="","",IF($P{row}>=Hypotheses_Marche!$B$7,"OK",IF($P{row}>=0,"A optimiser","Non rentable")))'
        marches[f"R{row}"] = f'=IF($A{row}="","",IF(AND($P{row}>=Hypotheses_Marche!$B$7,$O{row}>=1000),"Prioritaire",IF($P{row}<0,"A refuser/renegocier","A surveiller")))'
        marches[f"S{row}"] = f'=IF($A{row}="","",$H{row}-$N{row})'
        marches[f"T{row}"] = f'=IF($A{row}="","",IF($P{row}<0,"Renegocier prix/couts",IF($P{row}<Hypotheses_Marche!$B$7,"Optimiser heures/produits","Conserver")))'
        for col in range(1, 10):
            marches.cell(row, col).fill = PatternFill("solid", fgColor=INPUT)
        for col in range(10, 21):
            marches.cell(row, col).fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    marches.data_validations.dataValidation = []
    dv_type = DataValidation(type="list", formula1='"Jour,Nuit"', allow_blank=False)
    dv_prod = DataValidation(type="list", formula1='"Client,Nous"', allow_blank=False)
    marches.add_data_validation(dv_type)
    marches.add_data_validation(dv_prod)
    dv_type.add(f"E4:E{max_row}")
    dv_prod.add(f"F4:F{max_row}")
    tab = Table(displayName="MarchesRentabilite", ref=f"A3:T{max_row}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True)
    marches.add_table(tab)
    marches.freeze_panes = "A4"

    synth["A1"] = "SYNTHESE - RENTABILITE GLOBALE"
    rows = [
        ("Nombre de marches actifs", "=COUNT(Marches!H4:H25)"),
        ("CA total / mois (EUR)", "=SUM(Marches!H4:H25)"),
        ("Couts directs totaux / mois (EUR)", "=SUM(Marches!M4:M25)"),
        ("Marge de contribution totale / mois (EUR)", "=SUM(Marches!O4:O25)"),
        ("Charges fixes mensuelles (EUR)", "=Charges_Fixes!B14"),
        ("Resultat mensuel (EUR)", "=B7-B8"),
        ("Resultat annuel (EUR)", "=B9*12"),
        ("Marge de contribution moyenne %", "=IF(B5=0,0,B7/B5)"),
        ("Marge de contribution moyenne / marche (EUR)", "=IF(B4=0,0,B7/B4)"),
        ("Point mort (nb marches moyens p. couvrir charges fixes)", '=IF(B12<=0,"n/a",B8/B12)'),
        ("Marches rentables", '=COUNTIF(Marches!Q4:Q25,"OK")'),
        ("Marches a optimiser", '=COUNTIF(Marches!Q4:Q25,"A optimiser")'),
        ("Marches non rentables", '=COUNTIF(Marches!Q4:Q25,"Non rentable")'),
        ("Taux de marches rentables", "=IF(B4=0,0,B14/B4)"),
        ("Top marge mensuelle (EUR)", "=MAX(Marches!O4:O25)"),
    ]
    for idx, (label, formula) in enumerate(rows, 4):
        synth[f"A{idx}"] = label
        synth[f"B{idx}"] = formula

    if "Tableau_de_Bord" in wb.sheetnames:
        dash = wb["Tableau_de_Bord"]
        _clear(dash)
    else:
        dash = wb.create_sheet("Tableau_de_Bord", 0)
    dash["A1"] = "TABLEAU DE BORD FINANCIER CHRUTH"
    _title(dash["A1"])
    dash.merge_cells("A1:H1")
    dash["A2"] = "Vue de decision : rentabilite mensuelle, point mort, tresorerie previsionnelle et priorisation des marches."
    dash.merge_cells("A2:H2")
    dash["A2"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    kpis = [
        ("A4", "CA mensuel", "=Synthese_Rentabilite!B5"),
        ("C4", "Resultat mensuel", "=Synthese_Rentabilite!B9"),
        ("E4", "Marge moyenne", "=Synthese_Rentabilite!B11"),
        ("G4", "Point mort marches", "=Synthese_Rentabilite!B13"),
        ("A7", "Tresorerie fin annee", "=Previsionnel_Global!O11"),
        ("C7", "EBITDA annuel", "=Previsionnel_Global!O9"),
        ("E7", "Marches actifs", "=Synthese_Rentabilite!B4"),
        ("G7", "Taux marches rentables", "=Synthese_Rentabilite!B17"),
    ]
    for anchor, label, formula in kpis:
        cell = dash[anchor]
        cell.value = label
        cell.font = Font(bold=True, color=DARK)
        dash.cell(cell.row + 1, cell.column, formula).font = Font(bold=True, size=14, color=NAVY)
    dash["J3"], dash["K3"], dash["L3"], dash["M3"] = "Mois", "CA mensuel", "EBITDA", "Tresorerie cumulee"
    _header(dash["J3:M3"][0])
    for i, row in enumerate(range(4, 16), 26):
        dash[f"J{row}"] = f"=Previsionnel_Global!A{i}"
        dash[f"K{row}"] = f"=Previsionnel_Global!D{i}"
        dash[f"L{row}"] = f"=Previsionnel_Global!F{i}"
        dash[f"M{row}"] = f"=Previsionnel_Global!L{i}"
    line = LineChart()
    line.title = "CA, EBITDA et tresorerie previsionnels"
    line.add_data(Reference(dash, min_col=11, max_col=13, min_row=3, max_row=15), titles_from_data=True)
    line.set_categories(Reference(dash, min_col=10, min_row=4, max_row=15))
    dash.add_chart(line, "A16")

    if "Guide_Utilisation" in wb.sheetnames:
        guide = wb["Guide_Utilisation"]
        _clear(guide)
    else:
        guide = wb.create_sheet("Guide_Utilisation")
    guide["A1"] = "GUIDE DU MODELE FINANCIER CHRUTH"
    _title(guide["A1"])
    guide.merge_cells("A1:F1")
    guide["A3"], guide["B3"] = "Section", "Utilisation"
    _header(guide["A3:B3"][0])
    lines = [
        ("Objectif", "Piloter la rentabilite des marches de nettoyage et prioriser la prospection."),
        ("Previsionnel_Global", "Choisir un scenario et ajuster les hypotheses jaunes."),
        ("Marches", "Saisir les cellules jaunes ; les colonnes grises calculent couts, marge et decision."),
        ("Synthese_Rentabilite", "Lire resultat mensuel, point mort, taux de marches rentables et top marge."),
        ("Tableau_de_Bord", "Utiliser les KPI et graphiques pour presenter la decision business."),
        ("Version nettoyee", "Les logs, tests, caches et donnees brutes massives sont regenerables."),
    ]
    for idx, (section, usage) in enumerate(lines, 4):
        guide[f"A{idx}"] = section
        guide[f"B{idx}"] = usage
        guide[f"B{idx}"].alignment = Alignment(wrap_text=True)

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
    order = ["Tableau_de_Bord", "Previsionnel_Global", "Marches", "Synthese_Rentabilite", "Charges_Fixes", "Hypotheses_Marche", "Guide_Utilisation"]
    by_name = {ws.title: ws for ws in wb.worksheets}
    wb._sheets = [by_name[name] for name in order if name in by_name] + [ws for ws in wb.worksheets if ws.title not in order]
    wb.save(path)
    return path
