"""Verifie le cockpit produit (output/AO_CHRUTH.xlsm) contre les criteres de la spec."""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import ao_config  # noqa: E402

path = ao_config.AO_OUTPUT_XLSM
if not path.exists():
    print(f"FAIL fichier absent : {path}")
    raise SystemExit(1)

wb = openpyxl.load_workbook(path)
checks = []

checks.append(("Pilotage en 1er", wb.sheetnames[0] == "Pilotage"))
expected = [
    "Pilotage", "Top20_Ouverts", "AO_Tous", "AO_Ile_de_France", "AO_CHAUDS",
    "AO_A_VERIFIER", "DCE_A_RECUPERER", "CRM_Suivi", "Agent_IA", "Scoring", "Acheteurs",
    "Villes_Departements", "VBA_Bouton", "Logs_Update",
]
checks.append(("Tous onglets presents", all(s in wb.sheetnames for s in expected)))
checks.append(("Ordre des onglets exact", wb.sheetnames[: len(expected)] == expected))

pil = wb["Pilotage"]
flat = [str(c.value) for row in pil.iter_rows() for c in row if c.value is not None]
checks.append(("Bandeau PILOTAGE", any("PILOTAGE" in v.upper() for v in flat)))
checks.append((
    "7 KPI (date + 6 indicateurs)",
    all(any(lbl in v for v in flat) for lbl in
        ["Date MAJ", "Check qualite", "AO trouves", "AO chauds", "AO IDF", "Budget a verifier"]),
))

ao = wb["AO_Tous"]
header = {c.value: i for i, c in enumerate(ao[1], start=1)}
if ao.max_row > 1:
    qcell = ao[f"{get_column_letter(header['qualite_infos_formule'])}2"].value
    acell = ao[f"{get_column_letter(header['action_excel_formule'])}2"].value
    checks.append(("qualite_infos = formule", str(qcell).startswith("=")))
    checks.append(("action_excel = formule", str(acell).startswith("=")))
else:
    checks.append(("AO_Tous contient des donnees", False))

for c in ["resume_commercial", "proposition_message", "script_appel"]:
    checks.append((f"colonne IA {c} presente", c in header))

ok = all(v for _, v in checks)
for label, v in checks:
    print(("OK   " if v else "FAIL ") + label)
print("\nRESULTAT:", "TOUS OK" if ok else "ECHECS PRESENTS")
raise SystemExit(0 if ok else 1)
