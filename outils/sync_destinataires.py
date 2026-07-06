"""Lit la liste de destinataires saisie dans l'onglet Parametres du cockpit
(colonne dediee, une adresse par cellule) et la persiste dans destinataires.txt.

Appele par le bouton VBA "Enregistrer destinataires". Retour 0 si au moins une
adresse valide a ete ecrite, 1 sinon (pour le message VBA).
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import load_workbook  # noqa: E402

from ao_config import (  # noqa: E402
    ALERTE_DESTINATAIRES_FILE,
    ALERTE_SECRETS_FILE,
    AO_OUTPUT_XLSM,
    PARAM_DEST_COL,
    PARAM_DEST_FIRST_ROW,
    PARAM_DEST_MAX,
)


def lire_colonne(xlsm_path: Path = AO_OUTPUT_XLSM) -> list[str]:
    """Adresses valides (contenant '@') de la colonne destinataires, dedupliquees."""
    wb = load_workbook(xlsm_path, read_only=True, data_only=True)
    try:
        if "Parametres" not in wb.sheetnames:
            return []
        ws = wb["Parametres"]
        emails: list[str] = []
        for i in range(PARAM_DEST_MAX):
            cell = ws[f"{PARAM_DEST_COL}{PARAM_DEST_FIRST_ROW + i}"].value
            val = str(cell or "").strip()
            if val and "@" in val:
                emails.append(val)
        return list(dict.fromkeys(emails))
    finally:
        wb.close()


def ecrire_fichier(emails: list[str], dest_file: Path = ALERTE_DESTINATAIRES_FILE) -> None:
    entete = "# Destinataires des alertes CHRUTH (genere depuis l'onglet Parametres du cockpit).\n"
    Path(dest_file).write_text(entete + "\n".join(emails) + "\n", encoding="utf-8")


def ecrire_secrets(emails: list[str], secrets_file: Path = ALERTE_SECRETS_FILE) -> None:
    """Met a jour le champ destinataire sans toucher aux identifiants SMTP."""
    path = Path(secrets_file)
    data: dict[str, object] = {}
    if path.exists():
        raw = path.read_text(encoding="utf-8").strip()
        if raw:
            loaded = json.loads(raw)
            if not isinstance(loaded, dict):
                raise ValueError(f"Format invalide dans {path} : objet JSON attendu.")
            data = loaded
    data["destinataire"] = ", ".join(emails)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

def main() -> int:
    try:
        emails = lire_colonne()
    except FileNotFoundError:
        print("Cockpit introuvable :", AO_OUTPUT_XLSM)
        return 1
    if not emails:
        print("Aucune adresse valide dans la colonne destinataires.")
        return 1
    ecrire_fichier(emails)
    ecrire_secrets(emails)
    print(
        f"{len(emails)} destinataire(s) enregistre(s) -> "
        f"{ALERTE_DESTINATAIRES_FILE} + {ALERTE_SECRETS_FILE}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
