"""Simulateur de rentabilite par marche (par contrat de nettoyage) — CHRUTH.

Calcul bottom-up des couts d'un marche a partir de ses parametres operationnels
(jours/semaine, heures/jour, effectif, jour/nuit, produits, amortissement), puis
marge de CONTRIBUTION = prix - couts directs. Les charges fixes restent au niveau
entreprise (feuille Charges_Fixes) et la synthese fait : somme des marges - charges
fixes = resultat, + point mort en nombre de marches.

Les helpers Python ci-dessous sont le miroir des formules Excel ecrites par
ecrire_feuilles() ; ils servent aux tests et a d'eventuels calculs hors tableur.
Hypotheses (taux horaires, charges fixes, marches exemples) dans config.py.
"""
from __future__ import annotations

import config

from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation

_JAUNE = PatternFill("solid", fgColor="FFF2CC")
_GRAS = Font(bold=True)
_TITRE = Font(bold=True, size=14)
_EUR = "# ##0"
_PCT = "0%"
_H = "0.0"


def _feuille(wb, nom):
    """Recupere ou cree une feuille vide nommee `nom`."""
    if nom in wb.sheetnames:
        ws = wb[nom]
        if ws.max_row:
            ws.delete_rows(1, ws.max_row)
        return ws
    return wb.create_sheet(nom)


def heures_mois(jours_sem, heures_jour, effectif,
                semaines_mois=config.MARCHE_SEMAINES_PAR_MOIS) -> float:
    """Heures de nettoyage facturees par mois pour un marche recurrent."""
    return jours_sem * heures_jour * effectif * semaines_mois


def cout_main_oeuvre(h_mois, type_intervention,
                     taux_jour=config.MARCHE_TAUX_HORAIRE_JOUR,
                     taux_nuit=config.MARCHE_TAUX_HORAIRE_NUIT) -> float:
    """Cout main d'oeuvre mensuel : heures x taux horaire (majore la nuit)."""
    taux = taux_nuit if str(type_intervention).strip().lower().startswith("nuit") else taux_jour
    return h_mois * taux


def cout_produits(fournisseur, cout_produits_mois) -> float:
    """Cout produits mensuel : 0 si fournis par le client, sinon le montant saisi."""
    return cout_produits_mois if str(fournisseur).strip().lower() == "nous" else 0.0


def marge_marche(prix, cout_mo, cout_prod, amortissement) -> float:
    """Marge de contribution = prix - (main d'oeuvre + produits + amortissement)."""
    return prix - (cout_mo + cout_prod + amortissement)


def synthese(marches, charges_fixes_mensuelles) -> dict:
    """Agrege les marges des marches et soustrait les charges fixes entreprise.

    marches : liste de dicts avec au moins 'prix', 'cout_total', 'marge'.
    Renvoie les indicateurs de pilotage (resultat mensuel/annuel, marges, point
    mort = nombre de marches moyens necessaires pour couvrir les charges fixes).
    """
    nb = sum(1 for m in marches if m.get("prix"))
    ca = sum(m["prix"] for m in marches)
    couts = sum(m["cout_total"] for m in marches)
    marge_tot = sum(m["marge"] for m in marches)
    resultat = marge_tot - charges_fixes_mensuelles
    marge_moy_marche = (marge_tot / nb) if nb else 0.0
    point_mort = (charges_fixes_mensuelles / marge_moy_marche) if marge_moy_marche > 0 else None
    return {
        "nb_marches": nb,
        "ca_total": ca,
        "couts_directs": couts,
        "marge_contribution": marge_tot,
        "charges_fixes": charges_fixes_mensuelles,
        "resultat_mensuel": resultat,
        "resultat_annuel": resultat * 12,
        "marge_moyenne_pct": (marge_tot / ca) if ca else 0.0,
        "marge_moyenne_marche": marge_moy_marche,
        "point_mort_nb_marches": point_mort,
    }


def ecrire_feuilles(wb) -> None:
    """Ecrit les 4 feuilles du simulateur de rentabilite par marche dans `wb`
    (formules Excel, cellules jaunes editables). Tout se recalcule a l'edition."""
    # --- Feuille 1 : Charges_Fixes (entreprise, /mois) ---
    cf = _feuille(wb, "Charges_Fixes")
    cf["A1"] = "CHARGES FIXES ENTREPRISE (mensuelles)"
    cf["A1"].font = _TITRE
    cf["A2"] = ("Valeurs de test a ajuster. Les agents terrain et les produits "
                "specifiques a un marche sont sur la feuille Marches.")
    cf["A3"], cf["B3"] = "Charge fixe", "EUR / mois"
    cf["A3"].font = cf["B3"].font = _GRAS
    first = 4
    for i, (nom, val) in enumerate(config.MARCHE_CHARGES_FIXES.items()):
        r = first + i
        cf.cell(r, 1, nom)
        c = cf.cell(r, 2, val)
        c.number_format = _EUR
        c.fill = _JAUNE
    total_row = first + len(config.MARCHE_CHARGES_FIXES)   # 10 charges -> 14
    cf.cell(total_row, 1, "Total charges fixes").font = _GRAS
    tot = cf.cell(total_row, 2, f"=SUM(B{first}:B{total_row - 1})")
    tot.number_format = _EUR
    tot.font = _GRAS
    cf.column_dimensions["A"].width = 42
    cf.column_dimensions["B"].width = 14
    charges_fixes_ref = f"Charges_Fixes!B{total_row}"

    # --- Feuille 2 : Hypotheses_Marche (taux editables) ---
    hy = _feuille(wb, "Hypotheses_Marche")
    hy["A1"] = "HYPOTHESES DE CALCUL PAR MARCHE"
    hy["A1"].font = _TITRE
    hyp = [
        ("Cout horaire agent - JOUR (charge, EUR/h)", config.MARCHE_TAUX_HORAIRE_JOUR, _EUR),
        ("Cout horaire agent - NUIT (charge, EUR/h)", config.MARCHE_TAUX_HORAIRE_NUIT, _EUR),
        ("Semaines par mois", config.MARCHE_SEMAINES_PAR_MOIS, "0.00"),
        ("Cout produits / mois si fournis par nous (defaut EUR)", config.MARCHE_COUT_PRODUITS_DEFAUT, _EUR),
        ("Taux de marge cible (repere)", config.MARCHE_TAUX_MARGE_CIBLE, _PCT),
    ]
    for i, (label, val, fmt) in enumerate(hyp):
        r = 3 + i
        hy.cell(r, 1, label)
        c = hy.cell(r, 2, val)
        c.number_format = fmt
        c.fill = _JAUNE
    hy.column_dimensions["A"].width = 48
    TAUX_JOUR, TAUX_NUIT, SEM = "Hypotheses_Marche!$B$3", "Hypotheses_Marche!$B$4", "Hypotheses_Marche!$B$5"
    COUT_PROD_DEFAUT = "Hypotheses_Marche!$B$6"

    # --- Feuille 3 : Marches (1 ligne = 1 marche/client) ---
    ma = _feuille(wb, "Marches")
    ma["A1"] = "RENTABILITE PAR MARCHE — saisir les cellules jaunes, le reste se calcule"
    ma["A1"].font = _TITRE
    entete = 3
    headers = ["Nom du marche / client", "Jours/sem", "Heures/jour", "Effectif",
               "Type", "Produits par", "Cout produits/mois", "Prix du marche/mois",
               "Amortissement/mois", "Heures/mois", "Cout main d'oeuvre",
               "Cout produits", "Cout total", "Marge EUR", "Marge %"]
    for j, h in enumerate(headers, start=1):
        ma.cell(entete, j, h).font = _GRAS
    first = entete + 1                                        # 4
    n_data = len(config.MARCHE_EXEMPLES) + config.MARCHE_LIGNES_VIDES
    last = first + n_data - 1
    for i in range(n_data):
        r = first + i
        if i < len(config.MARCHE_EXEMPLES):
            nom, jours, heures, eff, typ, prod, cprod, prix, amort = config.MARCHE_EXEMPLES[i]
        else:
            nom, jours, heures, eff, typ, prod, cprod, prix, amort = ("", None, None, None, "Jour", "Nous", None, None, None)
        ma.cell(r, 1, nom)
        for col, val in ((2, jours), (3, heures), (4, eff), (5, typ),
                         (6, prod), (7, cprod), (8, prix), (9, amort)):
            c = ma.cell(r, col, val)
            c.fill = _JAUNE
        # colonnes calculees (formules)
        ma.cell(r, 10, f"=B{r}*C{r}*D{r}*{SEM}").number_format = _H            # heures/mois
        ma.cell(r, 11, f'=J{r}*IF(E{r}="Nuit",{TAUX_NUIT},{TAUX_JOUR})').number_format = _EUR  # MO
        # produits : pour un marche actif (prix saisi) fourni par nous, cout saisi
        # sinon le defaut (Hypotheses B6) ; 0 si fournis par le client ou ligne vide
        ma.cell(r, 12, f'=IF(AND(F{r}="Nous",H{r}<>""),IF(G{r}="",{COUT_PROD_DEFAUT},G{r}),0)').number_format = _EUR
        ma.cell(r, 13, f"=K{r}+L{r}+I{r}").number_format = _EUR               # cout total
        ma.cell(r, 14, f"=H{r}-M{r}").number_format = _EUR                    # marge EUR
        ma.cell(r, 15, f"=IF(H{r}=0,0,N{r}/H{r})").number_format = _PCT       # marge %
        for col in (7, 8, 9):
            ma.cell(r, col).number_format = _EUR
    # listes deroulantes Type (Jour/Nuit) et Produits (Nous/Client)
    dv_type = DataValidation(type="list", formula1='"Jour,Nuit"', allow_blank=True)
    dv_prod = DataValidation(type="list", formula1='"Nous,Client"', allow_blank=True)
    ma.add_data_validation(dv_type)
    ma.add_data_validation(dv_prod)
    dv_type.add(f"E{first}:E{last}")
    dv_prod.add(f"F{first}:F{last}")
    # marge negative en rouge
    rouge = Font(color="CC0000", bold=True)
    ma.conditional_formatting.add(
        f"N{first}:N{last}",
        CellIsRule(operator="lessThan", formula=["0"], font=rouge),
    )
    ma.column_dimensions["A"].width = 28
    for col in "BCDEFGHIJKLMNO":
        ma.column_dimensions[col].width = 13

    # --- Feuille 4 : Synthese_Rentabilite ---
    sy = _feuille(wb, "Synthese_Rentabilite")
    sy["A1"] = "SYNTHESE — RENTABILITE GLOBALE"
    sy["A1"].font = _TITRE
    lignes = [
        ("Nombre de marches actifs", f"=COUNT(Marches!H{first}:H{last})", "0"),
        ("CA total / mois (EUR)", f"=SUM(Marches!H{first}:H{last})", _EUR),
        ("Couts directs totaux / mois (EUR)", f"=SUM(Marches!M{first}:M{last})", _EUR),
        ("Marge de contribution totale / mois (EUR)", f"=SUM(Marches!N{first}:N{last})", _EUR),
        ("Charges fixes mensuelles (EUR)", f"={charges_fixes_ref}", _EUR),
        ("Resultat mensuel (EUR)", "=B7-B8", _EUR),
        ("Resultat annuel (EUR)", "=B9*12", _EUR),
        ("Marge de contribution moyenne %", "=IF(B5=0,0,B7/B5)", _PCT),
        ("Marge de contribution moyenne / marche (EUR)", "=IF(B4=0,0,B7/B4)", _EUR),
        ('Point mort (nb de marches moyens p. couvrir charges fixes)',
         '=IF(B12<=0,"n/a",B8/B12)', "0.0"),
    ]
    for i, (label, formule, fmt) in enumerate(lignes):
        r = 4 + i
        sy.cell(r, 1, label).font = _GRAS
        c = sy.cell(r, 2, formule)
        c.number_format = fmt
    sy.column_dimensions["A"].width = 52
    sy.column_dimensions["B"].width = 16
