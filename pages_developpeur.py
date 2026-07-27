"""Mode developpeur relie directement au cockpit Excel AO_CHRUTH.xlsm."""
from __future__ import annotations

import os
import subprocess
import sys
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

import livrables_chruth as livrables

try:
    st.set_page_config(page_title="Developpeur CHRUTH", layout="wide")
except st.errors.StreamlitAPIException:
    pass


@st.cache_data(show_spinner=False)
def structure_classeur(chemin: str, _modifie: int) -> list[dict[str, object]]:
    wb = load_workbook(chemin, read_only=True, data_only=True)
    try:
        return [{"Onglet": ws.title, "Lignes": ws.max_row, "Colonnes": ws.max_column}
                for ws in wb.worksheets]
    finally:
        wb.close()


@st.cache_data(show_spinner=False)
def apercu_onglet(chemin: str, _modifie: int, onglet: str, lignes: int) -> pd.DataFrame:
    return pd.read_excel(chemin, sheet_name=onglet, nrows=lignes, engine="openpyxl")


def ouvrir_windows(path: Path) -> None:
    if os.name != "nt":
        raise OSError("Ouverture locale disponible uniquement sous Windows.")
    os.startfile(str(path))  # type: ignore[attr-defined]


def lancer_mise_a_jour(script: Path) -> subprocess.Popen:
    logs = script.parent / "logs"
    logs.mkdir(parents=True, exist_ok=True)
    log_path = logs / "streamlit_developpeur_ao.log"
    flux = log_path.open("w", encoding="utf-8")
    creationflags = subprocess.CREATE_NO_WINDOW if os.name == "nt" else 0
    try:
        processus = subprocess.Popen(
            [sys.executable, str(script)], cwd=str(script.parent), stdout=flux,
            stderr=subprocess.STDOUT, creationflags=creationflags,
        )
    finally:
        flux.close()
    st.session_state["dev_log_path"] = str(log_path)
    return processus


st.title("Mode d\u00e9veloppeur")
st.caption("Acc\u00e8s technique au classeur Excel qui alimente le cockpit appels d'offres.")

actif = st.toggle("Activer le mode d\u00e9veloppeur", key="mode_developpeur_actif")
if not actif:
    st.info("Active ce mode pour inspecter les onglets, ouvrir Excel et lancer une mise \u00e0 jour.")
    st.stop()

classeur = livrables.fichier("AO_CHRUTH.xlsm")
if not classeur.exists():
    st.error(f"Classeur introuvable : {classeur}")
    st.stop()

info = livrables.informations(classeur)
verrou = livrables.verrou_excel(classeur)
script_update = livrables.racine_livraison() / "ao_weekly_update.py"

c1, c2, c3, c4 = st.columns(4)
c1.metric("Classeur", classeur.name)
c2.metric("Taille", livrables.taille_humaine(int(info["taille"])))
c3.metric("Mise \u00e0 jour", str(info["modifie"]))
c4.metric("\u00c9tat Excel", "Ouvert" if verrou.exists() else "Disponible")
st.code(str(classeur), language=None)
st.caption(f"Empreinte SHA-256 : `{livrables.empreinte(classeur)[:20]}...`")

if verrou.exists():
    st.warning("Excel semble ouvert. La lecture reste possible, mais ferme le classeur avant "
               "toute r\u00e9g\u00e9n\u00e9ration pour \u00e9viter un fichier verrouill\u00e9.")

b1, b2, b3 = st.columns(3)
if b1.button("Actualiser la liaison", type="primary"):
    st.cache_data.clear()
    st.rerun()
if b2.button("Ouvrir dans Excel", disabled=os.name != "nt"):
    try:
        ouvrir_windows(classeur)
        st.success("Classeur ouvert dans Excel.")
    except Exception as exc:  # noqa: BLE001
        st.error(f"Ouverture impossible : {exc}")
if b3.button("Ouvrir le dossier output", disabled=os.name != "nt"):
    try:
        ouvrir_windows(classeur.parent)
    except Exception as exc:  # noqa: BLE001
        st.error(f"Ouverture impossible : {exc}")

st.download_button("T\u00e9l\u00e9charger AO_CHRUTH.xlsm", classeur.read_bytes(),
                   file_name=classeur.name,
                   mime="application/vnd.ms-excel.sheet.macroEnabled.12")

st.subheader("Structure du classeur")
structure = structure_classeur(str(classeur), classeur.stat().st_mtime_ns)
st.dataframe(pd.DataFrame(structure), width="stretch", hide_index=True)

onglets = [str(ligne["Onglet"]) for ligne in structure]
c_onglet, c_lignes = st.columns([2, 1])
onglet = c_onglet.selectbox("Onglet \u00e0 pr\u00e9visualiser", onglets)
nb_lignes = c_lignes.select_slider("Lignes", options=[20, 50, 100, 200], value=50)
with st.spinner(f"Lecture de {onglet}..."):
    apercu = apercu_onglet(str(classeur), classeur.stat().st_mtime_ns, onglet, nb_lignes)
st.dataframe(apercu, width="stretch", height=470, hide_index=True)

st.subheader("Mise \u00e0 jour du cockpit Excel")
st.caption("Cette action lance le script de la m\u00eame livraison. Le classeur reste la source "
           "affich\u00e9e ci-dessus et ses macros sont pr\u00e9serv\u00e9es.")

if not script_update.exists():
    st.warning(f"Script de mise \u00e0 jour absent : {script_update}")
else:
    confirme = st.checkbox("Je confirme avoir ferm\u00e9 Excel avant la mise \u00e0 jour",
                           key="dev_confirme_excel_ferme")
    processus = st.session_state.get("dev_processus_ao")
    en_cours = processus is not None and processus.poll() is None
    if st.button("Lancer la mise \u00e0 jour AO", disabled=not confirme or verrou.exists() or en_cours):
        processus = lancer_mise_a_jour(script_update)
        st.session_state["dev_processus_ao"] = processus
        st.success(f"Mise \u00e0 jour lanc\u00e9e (processus {processus.pid}).")
        st.rerun()

    processus = st.session_state.get("dev_processus_ao")
    if processus is not None:
        code = processus.poll()
        if code is None:
            st.info(f"Mise \u00e0 jour en cours (processus {processus.pid}).")
        elif code == 0:
            st.success("Mise \u00e0 jour termin\u00e9e. Clique sur Actualiser la liaison.")
        else:
            st.error(f"La mise \u00e0 jour s'est termin\u00e9e avec le code {code}.")

    log_path = Path(st.session_state.get("dev_log_path", ""))
    if str(log_path) and log_path.is_file():
        lignes_log = log_path.read_text(encoding="utf-8", errors="replace").splitlines()
        with st.expander("Journal de la mise \u00e0 jour"):
            st.code("\n".join(lignes_log[-80:]) or "Journal vide", language=None)
