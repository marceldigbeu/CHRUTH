import pandas as pd
from openpyxl import load_workbook

import export_messages as em


def _frames():
    df_messages = pd.DataFrame([
        {"siret": "1", "denomination": "Soc1", "variante": "A",
         "template_id": "S|CHAUDE|A", "message_email": "Bonjour Soc1",
         "message_script": "Appel Soc1"},
    ])
    df_suivi = pd.DataFrame([
        {"siret": "1", "denomination": "Soc1", "ville": "Paris", "segment": "S|CHAUDE",
         "variante": "A", "template_id": "S|CHAUDE|A", "date_generation": "2026-06-25",
         "statut": "A_ENVOYER", "date_resultat": ""},
    ])
    df_perf = pd.DataFrame([
        {"segment": "S|CHAUDE", "variante": "A", "nb_envoyes": 0,
         "nb_resultats": 0, "taux_reponse": 0.0, "taux_rdv": 0.0},
    ])
    return df_messages, df_suivi, df_perf


def test_ecrire_classeur_cree_les_trois_feuilles(tmp_path):
    path = tmp_path / "msg.xlsx"
    dm, ds, dp = _frames()
    em.ecrire_classeur(dm, ds, dp, recommandations={}, path=path, seuil=20)
    wb = load_workbook(path)
    assert set(["Messages", "Suivi_Envois", "Perf_Messages"]).issubset(wb.sheetnames)


def test_suivi_a_une_validation_de_liste_sur_statut(tmp_path):
    path = tmp_path / "msg.xlsx"
    dm, ds, dp = _frames()
    em.ecrire_classeur(dm, ds, dp, recommandations={}, path=path, seuil=20)
    wb = load_workbook(path)
    ws = wb["Suivi_Envois"]
    assert len(ws.data_validations.dataValidation) >= 1
    formules = " ".join(dv.formula1 for dv in ws.data_validations.dataValidation)
    assert "RDV" in formules and "A_ENVOYER" in formules


def test_perf_contient_la_cellule_seuil(tmp_path):
    path = tmp_path / "msg.xlsx"
    dm, ds, dp = _frames()
    em.ecrire_classeur(dm, ds, dp, recommandations={"S|CHAUDE": "A"}, path=path, seuil=20)
    wb = load_workbook(path)
    ws = wb["Perf_Messages"]
    valeurs = [c.value for row in ws.iter_rows() for c in row]
    assert "Seuil" in valeurs and 20 in valeurs
