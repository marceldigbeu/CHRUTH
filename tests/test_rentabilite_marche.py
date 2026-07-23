import config
import rentabilite_marche as rm


def test_config_charges_fixes_complet():
    cf = config.MARCHE_CHARGES_FIXES
    assert isinstance(cf, dict)
    # les 10 charges nommees par Sonia
    attendus = [
        "Logiciel de paie", "Service Star", "Suite Google",
        "Assurance professionnelle - bureau", "Assurance professionnelle - Chruth",
        "Hebergement", "Internet", "Eco-pepiniere (loyer)",
        "Produits de nettoyage (stock structure)", "Salaires (structure / encadrement)",
    ]
    for nom in attendus:
        assert nom in cf, nom
    assert all(isinstance(v, (int, float)) for v in cf.values())


def test_config_hypotheses_et_exemples():
    assert config.MARCHE_TAUX_HORAIRE_NUIT > config.MARCHE_TAUX_HORAIRE_JOUR
    assert 4 <= config.MARCHE_SEMAINES_PAR_MOIS <= 5
    assert len(config.MARCHE_EXEMPLES) == 8
    # chaque exemple : 9 champs (nom, jours, heures, effectif, type, produits, cout_prod, prix, amort)
    for ex in config.MARCHE_EXEMPLES:
        assert len(ex) == 9


def test_heures_mois():
    # 5 j/sem x 3 h/j x 2 agents x 4.33 sem = 129.9 h/mois
    assert abs(rm.heures_mois(5, 3, 2, 4.33) - 129.9) < 0.01


def test_cout_main_oeuvre_jour_vs_nuit():
    h = 100
    jour = rm.cout_main_oeuvre(h, "Jour", taux_jour=18, taux_nuit=23)
    nuit = rm.cout_main_oeuvre(h, "Nuit", taux_jour=18, taux_nuit=23)
    assert jour == 1800
    assert nuit == 2300
    assert nuit > jour


def test_cout_produits_selon_fournisseur():
    assert rm.cout_produits("Nous", 60) == 60
    assert rm.cout_produits("Client", 60) == 0


def test_marge_marche_positive_et_negative():
    # prix 1800 ; couts = 1000 + 60 + 50 = 1110 -> marge 690
    assert rm.marge_marche(1800, 1000, 60, 50) == 690
    # marge negative possible
    assert rm.marge_marche(500, 800, 60, 50) < 0


def test_synthese_resultat_et_point_mort():
    marches = [
        {"prix": 1800, "cout_total": 1110, "marge": 690},
        {"prix": 4200, "cout_total": 3000, "marge": 1200},
    ]
    s = rm.synthese(marches, charges_fixes_mensuelles=3000)
    assert s["nb_marches"] == 2
    assert s["ca_total"] == 6000
    assert s["marge_contribution"] == 1890
    assert s["resultat_mensuel"] == 1890 - 3000  # -1110
    assert s["resultat_annuel"] == (1890 - 3000) * 12
    # marge moyenne par marche = 1890 / 2 = 945 ; point mort = 3000 / 945
    assert abs(s["point_mort_nb_marches"] - 3000 / 945) < 0.01


def test_synthese_point_mort_indefini_si_marge_negative():
    marches = [{"prix": 500, "cout_total": 800, "marge": -300}]
    s = rm.synthese(marches, charges_fixes_mensuelles=3000)
    assert s["point_mort_nb_marches"] is None


def test_ecrire_feuilles_cree_4_feuilles_avec_formules():
    from openpyxl import Workbook
    wb = Workbook()
    rm.ecrire_feuilles(wb)
    for nom in ("Charges_Fixes", "Hypotheses_Marche", "Marches", "Synthese_Rentabilite"):
        assert nom in wb.sheetnames, nom
    cf = wb["Charges_Fixes"]
    # total des charges fixes = SOMME (10 charges -> total en ligne 14)
    assert str(cf["B14"].value).startswith("=")
    assert "SUM" in str(cf["B14"].value).upper()
    # Marches : la marge (col N) de la 1re ligne de donnees est une formule
    mar = wb["Marches"]
    assert str(mar["N4"].value).startswith("=")
    # 8 exemples pre-remplis : le nom du 1er exemple est present
    assert mar["A4"].value == config.MARCHE_EXEMPLES[0][0]
    # Synthese : reference les charges fixes + a une ligne point mort
    syn = wb["Synthese_Rentabilite"]
    formules = [str(c.value) for row in syn.iter_rows() for c in row if c.value]
    assert any("Charges_Fixes!" in f for f in formules)
    labels = [str(c.value) for c in syn["A"]]
    assert any("Point mort" in str(x) for x in labels)
