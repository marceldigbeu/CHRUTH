import datetime

from openpyxl import Workbook

from ao_style import coerce_date, coerce_numeric, style_data_sheet


def test_style_adds_table_and_freeze():
    wb = Workbook()
    ws = wb.active
    ws.title = "AO_Tous"
    ws.append(["objet", "priorite", "budget_estime_eur"])
    ws.append(["Nettoyage", "CHAUD", 120000])
    ws.append(["Entretien", "FROID", 50000])
    style_data_sheet(ws, freeze_until="objet")

    assert ws.freeze_panes is not None
    assert len(ws.tables) == 1
    assert ws["A1"].font.color.rgb.endswith("FFFFFF")


def test_priority_color_and_safe_table_name():
    wb = Workbook()
    ws = wb.active
    ws.title = "AO Chauds & co"  # nom avec espaces et caractère invalide
    ws.append(["objet", "priorite"])
    ws.append(["X", "CHAUD"])
    style_data_sheet(ws)
    # couleur appliquée à la cellule priorite CHAUD
    assert ws["B2"].fill.start_color.rgb.endswith("C6EFCE")
    # nom de table sans espace ni esperluette
    tname = list(ws.tables.keys())[0]
    assert " " not in tname and "&" not in tname


def test_coerce_helpers():
    assert coerce_numeric("150000") == 150000.0
    assert coerce_numeric("150 000") == 150000.0
    assert coerce_numeric("") is None
    assert coerce_numeric("abc") is None
    assert coerce_date("2026-06-10") == datetime.date(2026, 6, 10)
    assert coerce_date("10/06/2026") == datetime.date(2026, 6, 10)
    assert coerce_date("2026-06-10T09:00:00") == datetime.date(2026, 6, 10)
    assert coerce_date("") is None


def test_number_and_date_formats_applied():
    wb = Workbook()
    ws = wb.active
    ws.title = "AO_Tous"
    ws.append(["objet", "budget_estime_eur", "date_limite"])
    ws.append(["Nettoyage", "150000", "2026-07-01"])
    ws.append(["Vide", "", ""])
    style_data_sheet(ws)
    # budget coerce en nombre + format euro
    assert ws["B2"].value == 150000.0
    assert "€" in ws["B2"].number_format
    # date coerce en date + format jj/mm/aaaa
    assert ws["C2"].value == datetime.date(2026, 7, 1)
    assert ws["C2"].number_format == "dd/mm/yyyy"
    # cellules vides laissees telles quelles (pas de crash)
    assert ws["B3"].value in ("", None)
