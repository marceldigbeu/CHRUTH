import pandas as pd
from openpyxl import load_workbook

import chruth_pipeline_master as master


class _FakeClient:
    def llm_disponible(self):
        return False  # repli deterministe, pas d'appel reseau

    def generer(self, *a, **k):
        return ""


def _chaude():
    return pd.DataFrame([
        {"siret": "1", "denomination": "Alpha", "libelle_commune": "Paris",
         "effectif_label": "10 a 19", "categorie_chruth": "PRIV_BUREAU",
         "priorite": "CHAUDE"},
        {"siret": "2", "denomination": "Beta", "libelle_commune": "Lyon",
         "effectif_label": "1 a 2", "categorie_chruth": "PRIV_BUREAU",
         "priorite": "CHAUDE"},
    ])


def test_generate_messages_ecrit_trois_feuilles_et_suivi(tmp_path):
    out = tmp_path / "Prospects_CHAUDS_messages.xlsx"
    suivi = tmp_path / "suivi.csv"
    df = master.generate_messages(
        source_df=_chaude(), suivi_path=suivi, output_path=out, client=_FakeClient())
    assert out.exists() and suivi.exists()
    wb = load_workbook(out)
    assert set(["Messages", "Suivi_Envois", "Perf_Messages"]).issubset(wb.sheetnames)
    # message rendu (placeholder remplace)
    assert "Alpha" in df.iloc[0]["message_email"]
    assert "{denomination}" not in df.iloc[0]["message_email"]
    # variantes alternees A/B sur les 2 prospects du meme segment
    assert set(df["variante"]) == {"A", "B"}


def test_generate_messages_preserve_statut_a_la_regeneration(tmp_path):
    out = tmp_path / "msg.xlsx"
    suivi = tmp_path / "suivi.csv"
    master.generate_messages(source_df=_chaude(), suivi_path=suivi,
                             output_path=out, client=_FakeClient())
    s = pd.read_csv(suivi, dtype=str)
    s.loc[s["siret"] == "1", "statut"] = "RDV"
    s.to_csv(suivi, index=False, encoding="utf-8")
    master.generate_messages(source_df=_chaude(), suivi_path=suivi,
                             output_path=out, client=_FakeClient())
    s2 = pd.read_csv(suivi, dtype=str)
    assert s2.loc[s2["siret"] == "1", "statut"].iloc[0] == "RDV"


class _SpyClient:
    """Records all LLM calls. Used to assert zero calls on the default (no-IA) path."""
    def __init__(self):
        self.calls: list[str] = []

    def llm_disponible(self):
        self.calls.append("llm_disponible")
        return False

    def generer(self, *a, **k):
        self.calls.append("generer")
        return ""


def test_generate_messages_default_zero_llm_calls_and_ab_distinct(tmp_path):
    """Default path (utiliser_ia=False by default): no LLM call, A and B texts differ."""
    out = tmp_path / "msg.xlsx"
    suivi = tmp_path / "suivi.csv"
    spy = _SpyClient()
    # Current code calls llm_disponible() once even when dispo=False; after fix: 0 calls.
    df = master.generate_messages(
        source_df=_chaude(), suivi_path=suivi, output_path=out, client=spy)
    assert spy.calls == [], (
        f"Default path (utiliser_ia=False) must make ZERO LLM calls, got: {spy.calls}")
    # Both variantes present (same segment, 2 prospects => A and B)
    assert set(df["variante"]) == {"A", "B"}
    msg_a = df.loc[df["variante"] == "A", "message_email"].iloc[0]
    msg_b = df.loc[df["variante"] == "B", "message_email"].iloc[0]
    # Deterministic A/B templates must be genuinely distinct
    assert msg_a != msg_b, "Default A and B emails must differ (deterministic templates)"
