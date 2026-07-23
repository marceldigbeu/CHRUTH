import previsions as pv


_P = {
    "horizon_mois": 12, "contacts_par_mois": 100, "taux_conversion": 0.03,
    "panier_annuel": 8000, "cout_variable_pct": 0.75,
    "couts_fixes_mensuels": 3000, "churn_mensuel": 0.0,
}


def test_projection_longueur_et_colonnes():
    proj = pv.projection(_P)
    assert len(proj) == 12
    for c in ["mois", "clients_actifs", "ca_mensuel_eur", "marge_brute_eur",
              "resultat_mensuel_eur", "resultat_cumule_eur"]:
        assert c in proj.columns


def test_clients_croissent_sans_churn():
    proj = pv.projection(_P)
    # sans churn : +3 clients/mois -> mois1=3, mois12=36
    assert round(proj["clients_actifs"].iloc[0]) == 3
    assert round(proj["clients_actifs"].iloc[-1]) == 36
    assert proj["clients_actifs"].iloc[-1] > proj["clients_actifs"].iloc[0]


def test_ca_coherent_avec_panier():
    proj = pv.projection(_P)
    # mois 1 : 3 clients x (8000/12) ~ 2000
    assert abs(proj["ca_mensuel_eur"].iloc[0] - 3 * 8000 / 12) < 1


def test_point_mort_detecte():
    proj = pv.projection(_P)
    pm = pv.point_mort_mois(proj)
    assert pm is None or (isinstance(pm, int) and 1 <= pm <= 12)


def test_point_mort_jamais_atteint():
    p = dict(_P, couts_fixes_mensuels=10_000_000)  # charges enormes
    assert pv.point_mort_mois(pv.projection(p)) is None


def test_ecrire_feuille_ecrit_des_formules():
    import config
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    pv.ecrire_feuille(ws)
    # selecteur de scenario + entree editable
    assert ws["B4"].value == config.PREV_SCENARIO_DEFAUT
    assert ws["B12"].value == config.PREV_CONTACTS_PAR_MOIS
    # conversion derivee du scenario (INDEX/MATCH)
    assert "INDEX" in str(ws["B13"].value)
    # projection (ligne 26) : clients / CA / EBITDA sont des formules
    for cell in ("C26", "D26", "F26", "J26", "L26"):
        assert str(ws[cell].value).startswith("=")
    # VAN et TRI presentes quelque part dans la feuille
    formules = [str(c.value) for row in ws.iter_rows() for c in row if c.value]
    assert any("NPV" in f for f in formules)
    assert any("IRR" in f for f in formules)


def test_hypotheses_synthese_contient_resultats():
    df = pv.hypotheses_synthese(_P)
    ind = set(df["Indicateur"])
    assert any("Point mort" in str(x) for x in ind)
    assert any("conversion" in str(x).lower() for x in ind)
    assert any("CA annee 1" in str(x) for x in ind)
