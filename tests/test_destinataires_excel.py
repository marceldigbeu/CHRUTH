import sys
from pathlib import Path

from openpyxl import Workbook

import ao_config
from ao_config import PARAM_DEST_FIRST_ROW

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "outils"))
import sync_destinataires as sd  # noqa: E402


def test_remplir_colonne_depuis_fichier(tmp_path, monkeypatch):
    f = tmp_path / "destinataires.txt"
    f.write_text("# entete ignoree\na@x.fr\n\npasunemail\nb@y.fr\n", encoding="utf-8")
    monkeypatch.setattr(ao_config, "ALERTE_DESTINATAIRES_FILE", f)
    from ao_export_excel import _remplir_destinataires

    wb = Workbook()
    ws = wb.active
    _remplir_destinataires(ws)
    assert ws[f"B{PARAM_DEST_FIRST_ROW}"].value == "a@x.fr"
    assert ws[f"B{PARAM_DEST_FIRST_ROW + 1}"].value == "b@y.fr"
    assert ws[f"B{PARAM_DEST_FIRST_ROW + 2}"].value is None


def test_sync_lit_colonne_filtre_et_dedup(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Parametres"
    ws[f"B{PARAM_DEST_FIRST_ROW}"] = "a@x.fr"
    ws[f"B{PARAM_DEST_FIRST_ROW + 1}"] = "pasunemail"
    ws[f"B{PARAM_DEST_FIRST_ROW + 2}"] = "b@y.fr"
    ws[f"B{PARAM_DEST_FIRST_ROW + 3}"] = "a@x.fr"  # doublon
    p = tmp_path / "cockpit.xlsx"
    wb.save(p)

    emails = sd.lire_colonne(p)
    assert emails == ["a@x.fr", "b@y.fr"]

    out = tmp_path / "dest.txt"
    sd.ecrire_fichier(emails, out)
    contenu = out.read_text(encoding="utf-8")
    assert "a@x.fr" in contenu and "b@y.fr" in contenu
    assert contenu.startswith("#")


def test_sync_sans_onglet_parametres(tmp_path):
    wb = Workbook()
    wb.active.title = "Autre"
    p = tmp_path / "cockpit.xlsx"
    wb.save(p)
    assert sd.lire_colonne(p) == []
