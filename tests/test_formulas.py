from openpyxl import Workbook

from ao_export_excel import DISPLAY_COLUMNS, add_excel_formulas


def _ws_with_headers():
    wb = Workbook()
    ws = wb.active
    for col, name in enumerate(DISPLAY_COLUMNS, start=1):
        ws.cell(row=1, column=col, value=name)
    for col in range(1, len(DISPLAY_COLUMNS) + 1):
        ws.cell(row=2, column=col, value="")
    return ws


def test_formulas_filled():
    ws = _ws_with_headers()
    headers = {c.value: i for i, c in enumerate(ws[1], start=1)}
    add_excel_formulas(ws, headers)
    from openpyxl.utils import get_column_letter
    qcol = get_column_letter(headers["qualite_infos_formule"])
    acol = get_column_letter(headers["action_excel_formule"])
    qf = ws[f"{qcol}2"].value
    af = ws[f"{acol}2"].value
    assert isinstance(qf, str) and qf.startswith("=")
    assert "Complet" in qf and "Partiel" in qf and "Vide" in qf
    assert isinstance(af, str) and af.startswith("=")
    assert "Telecharger DCE" in af and "Contacter maintenant" in af
