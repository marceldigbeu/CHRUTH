from datetime import date

import pandas as pd
from openpyxl import Workbook

from ao_pilotage import render_pilotage, compute_kpis
from ao_semaine import build_top20_ouverts


def _df():
    return pd.DataFrame(
        [
            {"id_ao": "b", "date_publication": "2026-06-10", "score_chruth": 95, "objet": "Nettoyage hopital",
             "acheteur": "CHRU", "ville": "Paris", "departement": "75", "departement_prestation": "",
             "budget_estime_eur": 150000, "date_limite": "2026-07-01", "url_avis": "http://x", "priorite": "CHAUD",
             "budget_statut": "", "statut_extraction": "OK"},
        ]
    )


def test_render_pilotage_structure():
    wb = Workbook()
    ws = wb.active
    ws.title = "Pilotage"
    df = _df()
    kpis = compute_kpis(df, today=date(2026, 6, 10))
    top20 = build_top20_ouverts(df, today=date(2026, 6, 10))
    render_pilotage(ws, kpis, top20)

    assert "PILOTAGE" in str(ws["A1"].value).upper()
    flat = [str(c.value) for row in ws.iter_rows() for c in row if c.value is not None]
    for label in ["Date MAJ", "Check qualite", "AO trouves", "AO chauds", "AO IDF", "Budget a verifier"]:
        assert any(label in v for v in flat), f"KPI manquant: {label}"
    assert any("Nettoyage hopital" in v for v in flat)
