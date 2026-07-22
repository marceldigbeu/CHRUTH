import zipfile

import pandas as pd

import scoring_export as se


def _df():
    return pd.DataFrame([
        {"siret": "1", "siren": "1", "denomination": "A", "categorie_chruth": "BUREAUX",
         "domaine_chruth": "PRIVE", "adresse_complete": "Paris", "libelle_commune": "Paris",
         "code_departement": "75", "signal_besoin": 80, "priorite": "CHAUDE",
         "latitude": 48.8, "longitude": 2.3},
        {"siret": "2", "siren": "2", "denomination": "B", "categorie_chruth": "COMMERCE",
         "domaine_chruth": "PRIVE", "adresse_complete": "Lyon", "libelle_commune": "Lyon",
         "code_departement": "69", "signal_besoin": 20, "priorite": "FROIDE",
         "latitude": 45.7, "longitude": 4.8},
    ])


def test_nom_fixe_xlsm():
    assert se.nom_fichier_sortie().name == "Base_Prospects_CHRUTH.xlsm"


def test_export_xlsm_garde_onglets_et_vba(tmp_path):
    out = tmp_path / "Base_Prospects_CHRUTH.xlsm"
    se.exporter_excel(_df(), out)  # la mise en forme est appliquee en memoire dans l'export
    import openpyxl
    wb = openpyxl.load_workbook(out, read_only=True)
    for s in ["Synthese", "Prospects", "CHAUDE", "Villes", "Societes"]:
        assert s in wb.sheetnames
    wb.close()
    assert any("vbaProject.bin" in n for n in zipfile.ZipFile(out).namelist())  # macro conservee


def _df_msg():
    return pd.DataFrame([
        {"siret": "1", "siren": "1", "denomination": "Alpha", "categorie_chruth": "PRIV_BUREAU",
         "domaine_chruth": "PRIVE", "adresse_complete": "Paris", "libelle_commune": "Paris",
         "code_departement": "75", "effectif_label": "10 a 19", "signal_besoin": 80,
         "priorite": "CHAUDE", "latitude": 48.8, "longitude": 2.3},
    ])


def test_export_avec_messages_onglet_et_colonnes(tmp_path, monkeypatch):
    # genere via repli deterministe (pas de LLM en test)
    monkeypatch.setenv("CHRUTH_GENERER_MESSAGES", "1")
    monkeypatch.setattr("prospect_messages.CACHE_PATH", tmp_path / "c.json")
    out = tmp_path / "Base_Prospects_CHRUTH.xlsm"
    df = _df_msg()
    df["signal_besoin"] = df.apply(se.calculer_signal_besoin, axis=1)
    df["priorite"] = df["signal_besoin"].apply(se.priorite)
    se.exporter_excel(df, out)
    import openpyxl
    wb = openpyxl.load_workbook(out, read_only=True)
    assert "Templates_Segments" in wb.sheetnames
    entetes_chaude = [c.value for c in next(wb["CHAUDE"].iter_rows(max_row=1))]
    assert "brouillon_email" in entetes_chaude
    # spec : PAS de colonne brouillon sur la grosse feuille Prospects (132k)
    entetes_prospects = [c.value for c in next(wb["Prospects"].iter_rows(max_row=1))]
    assert "brouillon_email" not in entetes_prospects
    wb.close()
