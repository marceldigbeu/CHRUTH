from functools import partial

import openpyxl
import pytest

import ao_config
from ao_db import fetch_logs, fetch_records, upsert_records
from ao_export_excel import export_excel


@pytest.mark.skipif(not ao_config.AO_TEMPLATE_XLSM.exists(), reason="gabarit requis (Task 3)")
def test_export_produces_xlsm_with_pilotage_first(tmp_path):
    # Ecrit dans un fichier temporaire (jamais le vrai output/AO_CHRUTH.xlsm,
    # qui peut etre verrouille par Excel ouvert) -> test isole et robuste.
    path = export_excel(path=tmp_path / "AO_CHRUTH.xlsm")
    assert path.suffix == ".xlsm"
    wb = openpyxl.load_workbook(path, read_only=True)
    assert wb.sheetnames[0] == "Pilotage"
    for expected in ["Top20_Ouverts", "AO_Tous", "AO_CHAUDS", "CRM_Suivi", "Agent_IA", "Scoring", "Logs_Update"]:
        assert expected in wb.sheetnames
    wb.close()


@pytest.mark.skipif(not ao_config.AO_TEMPLATE_XLSM.exists(), reason="gabarit requis (Task 3)")
def test_export_essentiels_sheet_and_full_detail(tmp_path, monkeypatch):
    """AO_Nettoyage_IDF (essentiels) doit exister sans rien retirer de AO_Tous."""
    db_path = tmp_path / "ao_test.sqlite"
    upsert_records(
        [
            {
                "id_ao": "AO1",
                "objet": "Nettoyage des locaux scolaires",
                "acheteur": "Mairie de Paris",
                "secteur": "Ecole",
                "categorie": "BATIMENTS",
                "ville": "Paris",
                "departement": "75",
                "departement_prestation": "75",
                "date_publication": "2026-06-01",
                "date_limite": "2026-07-01",
                "budget_estime_eur": "80000",
                "budget_annuel_eur": "80000",
                "budget_annualise": "OUI",
                "url_dce": "https://example.test/dce/AO1",
                "url_avis": "https://example.test/avis/AO1",
                "priorite": "CHAUD",
                "score_chruth": "75",
            }
        ],
        db_path=db_path,
    )

    monkeypatch.setattr("ao_export_excel.fetch_records", partial(fetch_records, db_path=db_path))
    monkeypatch.setattr("ao_export_excel.fetch_logs", partial(fetch_logs, db_path=db_path))

    out = tmp_path / "AO_CHRUTH.xlsm"
    path = export_excel(path=out)

    wb = openpyxl.load_workbook(path, read_only=True)
    assert "AO_Nettoyage_IDF" in wb.sheetnames

    essentiels = wb["AO_Nettoyage_IDF"]
    headers = [c.value for c in next(essentiels.iter_rows(min_row=1, max_row=1))]
    assert "Catégorie" in headers
    assert "Lien DCE" in headers

    assert "AO_Tous" in wb.sheetnames
    ao_tous = wb["AO_Tous"]
    ao_tous_headers = [c.value for c in next(ao_tous.iter_rows(min_row=1, max_row=1))]
    assert "raisons_scoring" in ao_tous_headers
    wb.close()
